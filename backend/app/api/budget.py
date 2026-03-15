"""Budget model API — trigger calculation, check status, read outputs."""

from __future__ import annotations

import io
import uuid
from collections import defaultdict
from datetime import datetime, timezone

from celery.result import AsyncResult
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_db, require_finance
from app.db.models.account import Account, AccountType, Statement
from app.db.models.budget import BudgetVersion, ModelAssumption, ModelOutput
from app.db.models.debt import DebtFacility, DebtSchedule
from app.db.models.entity import Entity
from app.db.models.location import Location, SiteBudgetEntry
from app.db.models.period import Period, WeeklyPeriod
from app.db.models.user import User
from app.db.models.wc import WcDriver
from app.schemas.budget import (
    AssumptionPayload,
    BudgetVersionCreate,
    BudgetVersionRead,
    CalculationStatusResponse,
    CalculationTriggerResponse,
    DebtFacilityRead,
    DebtFacilityUpdate,
    DebtHistoryPoint,
    DebtScheduleRowRead,
    DebtSummary,
    ModelAssumptionRead,
    ModelOutputResponse,
    ModelOutputRow,
    SiteAnnualSummaryRow,
    SiteBudgetAssumptionBulkUpdate,
    SiteBudgetAssumptionRead,
    SiteBudgetAssumptionUpdate,
    SiteBudgetGridRead,
    SiteBudgetLineRead,
    SiteBudgetSavePayload,
    SiteRollupRow,
    SiteSummaryRead,
    SiteWeeklyBudgetRow,
    SiteWeeklyOverridePayload,
    WcDriverRead,
    WcDriverUpdate,
)

router = APIRouter(prefix="/budgets", tags=["budgets"])

MONTH_ABBR = [
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
]


def _period_label(period: Period) -> str:
    if period.period_start:
        return period.period_start.strftime("%b-%y")
    cal_year = period.fy_year - 1 if period.fy_month <= 6 else period.fy_year
    return f"{MONTH_ABBR[period.fy_month - 1]}-{cal_year % 100:02d}"


# ── Section header config ────────────────────────────────────────────────────

IS_SECTIONS: dict[str, list[AccountType]] = {
    "Revenue": [AccountType.income, AccountType.cogs],
    "Operating Expenses": [AccountType.opex],
    "Depreciation & Amortisation": [AccountType.depreciation],
    "Interest": [AccountType.interest],
    "Tax": [AccountType.tax],
}

BS_SECTIONS: dict[str, list[AccountType]] = {
    "Assets": [AccountType.asset],
    "Liabilities": [AccountType.liability],
    "Equity": [AccountType.equity],
}

CF_SECTIONS: dict[str, list[str]] = {
    "Operating Activities": ["CF-OPERATING"],
    "Investing Activities": ["CF-INVESTING"],
    "Financing Activities": ["CF-FINANCING"],
    "Net Cash Flow": ["CF-NET"],
}


# ── GET /budgets ──────────────────────────────────────────────────────────────


@router.get("/", response_model=list[BudgetVersionRead])
async def list_budget_versions(
    fy_year: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    q = select(BudgetVersion).order_by(BudgetVersion.created_at.desc())
    if fy_year is not None:
        q = q.where(BudgetVersion.fy_year == fy_year)
    result = await db.execute(q)
    return list(result.scalars().all())


# ── POST /budgets ─────────────────────────────────────────────────────────────


@router.post("/", response_model=BudgetVersionRead, status_code=status.HTTP_201_CREATED)
async def create_budget_version(
    payload: BudgetVersionCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_finance),
):
    version = BudgetVersion(
        name=payload.name,
        fy_year=payload.fy_year,
        version_type=payload.version_type,
        base_version_id=payload.base_version_id,
        created_by=user.id,
    )
    db.add(version)
    await db.commit()
    await db.refresh(version)
    return version


# ── GET /budgets/{id}/assumptions ─────────────────────────────────────────────


@router.get("/{budget_id}/assumptions", response_model=list[ModelAssumptionRead])
async def get_assumptions(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(ModelAssumption).where(
            ModelAssumption.budget_version_id == budget_id
        )
    )
    return list(result.scalars().all())


# ── PUT /budgets/{id}/assumptions ─────────────────────────────────────────────


@router.put("/{budget_id}/assumptions", response_model=list[ModelAssumptionRead])
async def save_assumptions(
    budget_id: uuid.UUID,
    assumptions: list[AssumptionPayload],
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    for payload in assumptions:
        result = await db.execute(
            select(ModelAssumption).where(
                ModelAssumption.budget_version_id == budget_id,
                ModelAssumption.entity_id == payload.entity_id,
                ModelAssumption.location_id == payload.location_id,
                ModelAssumption.assumption_key == payload.assumption_key,
            )
        )
        existing = result.scalar_one_or_none()

        if existing:
            existing.assumption_value = payload.assumption_value
            existing.updated_by = user.id
            existing.updated_at = datetime.now(timezone.utc)
        else:
            db.add(ModelAssumption(
                budget_version_id=budget_id,
                entity_id=payload.entity_id,
                location_id=payload.location_id,
                assumption_key=payload.assumption_key,
                assumption_value=payload.assumption_value,
                updated_by=user.id,
            ))

    await db.commit()

    result = await db.execute(
        select(ModelAssumption).where(
            ModelAssumption.budget_version_id == budget_id
        )
    )
    return list(result.scalars().all())


# ── GET /budgets/{id}/wc-drivers ──────────────────────────────────────────────


@router.get("/{budget_id}/wc-drivers", response_model=list[WcDriverRead])
async def get_wc_drivers(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(WcDriver).where(WcDriver.budget_version_id == budget_id)
    )
    drivers = list(result.scalars().all())

    account_ids = {d.account_id for d in drivers}
    acct_result = await db.execute(
        select(Account).where(Account.id.in_(account_ids))
    ) if account_ids else None
    acct_map = {a.id: a.name for a in acct_result.scalars().all()} if acct_result else {}

    return [
        WcDriverRead(
            id=d.id,
            entity_id=d.entity_id,
            account_id=d.account_id,
            account_label=acct_map.get(d.account_id, ""),
            driver_type=d.driver_type.value if d.driver_type else None,
            base_days=float(d.base_days) if d.base_days is not None else None,
            seasonal_factors=d.seasonal_factors,
            notes=d.notes,
        )
        for d in drivers
    ]


# ── PUT /budgets/{id}/wc-drivers ──────────────────────────────────────────────


@router.put("/{budget_id}/wc-drivers", response_model=list[WcDriverRead])
async def update_wc_drivers(
    budget_id: uuid.UUID,
    updates: list[WcDriverUpdate],
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    for upd in updates:
        driver = await db.get(WcDriver, upd.id)
        if driver and driver.budget_version_id == budget_id:
            driver.base_days = upd.base_days
            driver.seasonal_factors = upd.seasonal_factors
            driver.last_updated_by = user.id
            driver.last_updated_at = datetime.now(timezone.utc)

    await db.commit()
    return await get_wc_drivers(budget_id, db, user)


# ── Auto-seed debt facilities from BS-DEBT-* accounts ────────────────────────


def _infer_facility_type(code: str, name: str) -> str | None:
    code_lower = code.lower()
    name_lower = name.lower()
    if "equip" in code_lower or "equip" in name_lower:
        return "equipment_loan"
    if "vehicle" in code_lower or "vehicle" in name_lower:
        return "vehicle_loan"
    return "property_loan"


async def _auto_seed_debt_facilities(
    db: AsyncSession,
) -> list[DebtFacility]:
    """Discover BS-DEBT-* accounts and create DebtFacility records if none exist."""
    from app.db.models.consolidation import ConsolidatedActual

    existing = await db.execute(select(DebtFacility))
    if list(existing.scalars().all()):
        return []

    acct_result = await db.execute(
        select(Account).where(
            Account.code.like("BS-DEBT-%"),
            Account.is_subtotal.is_(False),
        ).order_by(Account.sort_order)
    )
    debt_accounts = list(acct_result.scalars().all())
    if not debt_accounts:
        return []

    entity_result = await db.execute(
        select(Entity).where(Entity.is_active.is_(True)).order_by(Entity.code)
    )
    entities = list(entity_result.scalars().all())
    entity_map = {e.id: e for e in entities}
    default_entity_id = entities[0].id if entities else None

    created = []
    for idx, acct in enumerate(debt_accounts):
        balance_result = await db.execute(
            select(
                ConsolidatedActual.entity_id,
                ConsolidatedActual.amount,
            )
            .where(
                ConsolidatedActual.account_id == acct.id,
                ConsolidatedActual.is_group_total.is_(False),
                ConsolidatedActual.entity_id.isnot(None),
            )
            .order_by(ConsolidatedActual.calculated_at.desc())
        )
        balance_rows = balance_result.all()

        entity_balances: dict[uuid.UUID, float] = {}
        for row in balance_rows:
            if row.entity_id not in entity_balances:
                entity_balances[row.entity_id] = abs(float(row.amount or 0))

        if entity_balances:
            best_entity = max(entity_balances, key=lambda e: entity_balances[e])
            opening = max(entity_balances.values())
        else:
            best_entity = default_entity_id
            opening = 0.0

        if best_entity is None:
            continue

        fac = DebtFacility(
            code=acct.code,
            name=acct.name,
            entity_id=best_entity,
            facility_type=_infer_facility_type(acct.code, acct.name),
            opening_balance=opening,
            base_rate=None,
            margin=0,
            monthly_repayment=None,
            sort_order=idx,
            is_active=True,
        )
        db.add(fac)
        created.append(fac)

    if created:
        await db.flush()

    return created


# ── GET /budgets/{id}/debt ────────────────────────────────────────────────────


@router.get("/{budget_id}/debt", response_model=DebtSummary)
async def get_debt_facilities(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    from app.db.models.consolidation import ConsolidatedActual

    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    fac_result = await db.execute(
        select(DebtFacility)
        .where(DebtFacility.is_active.is_(True))
        .order_by(DebtFacility.sort_order)
    )
    facilities = list(fac_result.scalars().all())

    if not facilities:
        seeded = await _auto_seed_debt_facilities(db)
        if seeded:
            await db.commit()
            facilities = seeded

    entity_result = await db.execute(select(Entity))
    entity_map = {e.id: e.code for e in entity_result.scalars().all()}

    # Load all periods across FYs for history
    all_period_result = await db.execute(
        select(Period).order_by(Period.fy_year, Period.fy_month)
    )
    all_periods = list(all_period_result.scalars().all())
    period_map = {p.id: p for p in all_periods}

    # Load FY periods for schedule
    fy_period_result = await db.execute(
        select(Period)
        .where(Period.fy_year == version.fy_year)
        .order_by(Period.fy_month)
    )
    fy_periods = list(fy_period_result.scalars().all())
    fy_period_ids = [p.id for p in fy_periods]

    # Load schedules
    sched_result = await db.execute(
        select(DebtSchedule).where(
            DebtSchedule.budget_version_id == budget_id,
            DebtSchedule.period_id.in_(fy_period_ids),
        )
    ) if fy_period_ids else None
    schedules = list(sched_result.scalars().all()) if sched_result else []
    sched_by_fac: dict[uuid.UUID, list] = defaultdict(list)
    for s in schedules:
        sched_by_fac[s.facility_id].append(s)

    # Load historical balances from consolidated_actuals for BS-DEBT-* accounts
    acct_code_result = await db.execute(
        select(Account).where(
            Account.code.like("BS-DEBT-%"),
            Account.is_subtotal.is_(False),
        )
    )
    debt_accounts = list(acct_code_result.scalars().all())
    acct_code_to_id = {a.code: a.id for a in debt_accounts}
    acct_id_to_code = {a.id: a.code for a in debt_accounts}

    all_debt_acct_ids = list(acct_code_to_id.values())
    history_result = await db.execute(
        select(ConsolidatedActual).where(
            ConsolidatedActual.account_id.in_(all_debt_acct_ids),
            ConsolidatedActual.is_group_total.is_(True),
        )
    ) if all_debt_acct_ids else None
    all_history = list(history_result.scalars().all()) if history_result else []

    # Group history by account code
    history_by_code: dict[str, dict[uuid.UUID, float]] = defaultdict(dict)
    for h in all_history:
        code = acct_id_to_code.get(h.account_id, "")
        if code:
            history_by_code[code][h.period_id] = abs(float(h.amount or 0))

    # Build total debt history
    total_by_period: dict[uuid.UUID, float] = defaultdict(float)
    for code, period_vals in history_by_code.items():
        for pid, val in period_vals.items():
            total_by_period[pid] += val

    total_debt_history: list[DebtHistoryPoint] = []
    for p in all_periods:
        if p.id in total_by_period:
            total_debt_history.append(DebtHistoryPoint(
                period_label=_period_label(p),
                fy_year=p.fy_year,
                fy_month=p.fy_month,
                balance=total_by_period[p.id],
                movement=0,
            ))

    # Calculate movements for total
    for i in range(1, len(total_debt_history)):
        total_debt_history[i].movement = (
            total_debt_history[i].balance - total_debt_history[i - 1].balance
        )

    period_label_map = {p.id: _period_label(p) for p in fy_periods}

    total_interest_budget = 0.0
    total_repayment_budget = 0.0

    result_facilities = []
    for fac in facilities:
        # Build schedule rows
        fac_schedules = sorted(
            sched_by_fac.get(fac.id, []),
            key=lambda s: fy_period_ids.index(s.period_id) if s.period_id in fy_period_ids else 999,
        )
        schedule_rows = [
            DebtScheduleRowRead(
                period_label=period_label_map.get(s.period_id, ""),
                opening_balance=float(s.opening_balance or 0),
                drawdown=float(s.drawdown or 0),
                repayment=float(s.repayment or 0),
                closing_balance=float(s.closing_balance or 0),
                interest_expense=float(s.interest_expense or 0),
                interest_rate_applied=float(s.interest_rate_applied or 0),
            )
            for s in fac_schedules
        ]

        for sr in fac_schedules:
            total_interest_budget += float(sr.interest_expense or 0)
            total_repayment_budget += float(sr.repayment or 0)

        # Build history for this facility
        fac_history_data = history_by_code.get(fac.code, {})
        history_points: list[DebtHistoryPoint] = []
        for p in all_periods:
            if p.id in fac_history_data:
                history_points.append(DebtHistoryPoint(
                    period_label=_period_label(p),
                    fy_year=p.fy_year,
                    fy_month=p.fy_month,
                    balance=fac_history_data[p.id],
                    movement=0,
                ))

        # Calculate month-over-month movements and implied amortization
        movements = []
        for i in range(1, len(history_points)):
            mv = history_points[i].balance - history_points[i - 1].balance
            history_points[i].movement = mv
            if mv < 0:
                movements.append(abs(mv))

        avg_monthly_repayment = None
        if movements:
            avg_monthly_repayment = sum(movements) / len(movements)

        # Implied interest rate (from average balance × observed interest if available)
        implied_rate = None
        if fac.base_rate is not None:
            implied_rate = float(fac.base_rate or 0) + float(fac.margin or 0)

        current_balance = history_points[-1].balance if history_points else float(fac.opening_balance)

        result_facilities.append(DebtFacilityRead(
            id=fac.id,
            code=fac.code,
            name=fac.name,
            entity_id=fac.entity_id,
            entity_code=entity_map.get(fac.entity_id),
            facility_type=fac.facility_type.value if fac.facility_type else None,
            opening_balance=float(fac.opening_balance),
            current_balance=current_balance,
            base_rate=float(fac.base_rate) if fac.base_rate is not None else None,
            margin=float(fac.margin),
            amort_type=fac.amort_type.value if fac.amort_type else None,
            monthly_repayment=float(fac.monthly_repayment) if fac.monthly_repayment is not None else None,
            maturity_date=str(fac.maturity_date) if fac.maturity_date else None,
            is_active=fac.is_active,
            schedule=schedule_rows,
            history=history_points,
            implied_interest_rate=implied_rate,
            avg_monthly_repayment=avg_monthly_repayment,
        ))

    total_debt = sum(f.current_balance or f.opening_balance for f in result_facilities)

    return DebtSummary(
        total_debt=total_debt,
        total_interest_budget=total_interest_budget,
        total_repayment_budget=total_repayment_budget,
        facility_count=len(result_facilities),
        facilities=result_facilities,
        total_debt_history=total_debt_history,
    )


# ── PUT /budgets/{id}/debt/facilities/{fid} ──────────────────────────────────


@router.put("/{budget_id}/debt/facilities/{facility_id}", response_model=DebtFacilityRead)
async def update_debt_facility(
    budget_id: uuid.UUID,
    facility_id: uuid.UUID,
    payload: DebtFacilityUpdate,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    facility = await db.get(DebtFacility, facility_id)
    if facility is None:
        raise HTTPException(status_code=404, detail="Facility not found")

    facility.base_rate = payload.base_rate
    facility.margin = payload.margin
    facility.monthly_repayment = payload.monthly_repayment
    await db.commit()
    await db.refresh(facility)

    entity_result = await db.execute(select(Entity))
    entity_map = {e.id: e.code for e in entity_result.scalars().all()}

    return DebtFacilityRead(
        id=facility.id,
        code=facility.code,
        name=facility.name,
        entity_id=facility.entity_id,
        entity_code=entity_map.get(facility.entity_id),
        facility_type=facility.facility_type.value if facility.facility_type else None,
        opening_balance=float(facility.opening_balance),
        base_rate=float(facility.base_rate) if facility.base_rate is not None else None,
        margin=float(facility.margin),
        amort_type=facility.amort_type.value if facility.amort_type else None,
        monthly_repayment=float(facility.monthly_repayment) if facility.monthly_repayment is not None else None,
        maturity_date=str(facility.maturity_date) if facility.maturity_date else None,
        is_active=facility.is_active,
        schedule=[],
    )


# ── GET /budgets/{id}/sites ───────────────────────────────────────────────────


@router.get("/{budget_id}/sites", response_model=list[SiteSummaryRead])
async def list_sites(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(Location)
        .where(Location.is_active.is_(True))
        .order_by(Location.state, Location.name)
    )
    locations = list(result.scalars().all())

    # Load periods for labels
    period_result = await db.execute(
        select(Period)
        .where(Period.fy_year == version.fy_year)
        .order_by(Period.fy_month)
    )
    periods = list(period_result.scalars().all())
    period_labels = {p.fy_month: _period_label(p) for p in periods}

    # Load weekly periods for proration
    wp_result = await db.execute(
        select(WeeklyPeriod).where(WeeklyPeriod.fy_year == version.fy_year)
    )
    weekly_periods = list(wp_result.scalars().all())
    week_to_month: dict[uuid.UUID, int] = {}
    week_proration: dict[uuid.UUID, float] = {}
    for wp in weekly_periods:
        if wp.fy_month is not None and wp.days_this_week_in_fy_month is not None:
            week_to_month[wp.id] = wp.fy_month
            week_proration[wp.id] = wp.days_this_week_in_fy_month / 7.0

    # Load all site budget entries for this version
    entry_result = await db.execute(
        select(SiteBudgetEntry).where(SiteBudgetEntry.version_id == budget_id)
    )
    entries = list(entry_result.scalars().all())

    # Aggregate revenue per site per month
    site_month_totals: dict[uuid.UUID, dict[str, float]] = defaultdict(
        lambda: defaultdict(float)
    )
    revenue_lines = {"boarding_revenue", "grooming_revenue", "other_revenue"}
    for entry in entries:
        if entry.model_line_item not in revenue_lines:
            continue
        if entry.week_id and entry.week_id in week_to_month:
            fy_month = week_to_month[entry.week_id]
            factor = week_proration.get(entry.week_id, 1.0)
            amount = float(entry.amount or 0) * factor
        else:
            continue
        label = period_labels.get(fy_month, f"M{fy_month:02d}")
        site_month_totals[entry.location_id][label] += amount

    return [
        SiteSummaryRead(
            location_id=loc.id,
            code=loc.code or "",
            name=loc.name or "",
            state=loc.state,
            entity_id=loc.entity_id,
            capacity_dogs=loc.capacity_dogs,
            monthly_totals=dict(site_month_totals.get(loc.id, {})),
        )
        for loc in locations
    ]


# ── GET /budgets/{id}/sites/annual-summary ────────────────────────────────────


@router.get(
    "/{budget_id}/sites/annual-summary",
    response_model=list[SiteAnnualSummaryRow],
)
async def get_sites_annual_summary(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    """All-sites annual summary for a budget version."""
    from sqlalchemy import func as sqla_func

    from app.db.models.location import SiteBudgetAssumption, SiteWeeklyBudget

    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(Location)
        .where(Location.is_active.is_(True))
        .order_by(Location.state, Location.name)
    )
    locations = result.scalars().all()

    result = await db.execute(
        select(
            SiteWeeklyBudget.location_id,
            sqla_func.sum(SiteWeeklyBudget.prior_year_pet_days_boarding + SiteWeeklyBudget.prior_year_pet_days_daycare).label("prior_pd"),
            sqla_func.sum(SiteWeeklyBudget.budget_pet_days_boarding + SiteWeeklyBudget.budget_pet_days_daycare).label("budget_pd"),
            sqla_func.sum(SiteWeeklyBudget.prior_year_revenue).label("prior_rev"),
            sqla_func.sum(SiteWeeklyBudget.budget_revenue).label("budget_rev"),
            sqla_func.sum(SiteWeeklyBudget.budget_labour).label("budget_lab"),
            sqla_func.sum(
                SiteWeeklyBudget.budget_cogs
                + SiteWeeklyBudget.budget_rent
                + SiteWeeklyBudget.budget_utilities
                + SiteWeeklyBudget.budget_rm
                + SiteWeeklyBudget.budget_it
                + SiteWeeklyBudget.budget_general
                + SiteWeeklyBudget.budget_advertising
            ).label("budget_costs"),
        )
        .where(SiteWeeklyBudget.version_id == budget_id)
        .group_by(SiteWeeklyBudget.location_id)
    )
    agg_rows = {r.location_id: r for r in result.all()}

    result = await db.execute(
        select(SiteBudgetAssumption).where(
            SiteBudgetAssumption.version_id == budget_id
        )
    )
    assumptions_map = {a.location_id: a for a in result.scalars().all()}

    output = []
    for loc in locations:
        agg = agg_rows.get(loc.id)
        assumption = assumptions_map.get(loc.id)

        if assumption:
            a_status = "locked" if assumption.assumptions_locked else "set"
        else:
            a_status = "default" if agg else "no_data"

        prior_pd = int(agg.prior_pd or 0) if agg else 0
        budget_pd = int(agg.budget_pd or 0) if agg else 0
        prior_rev = float(agg.prior_rev or 0) if agg else 0
        budget_rev = float(agg.budget_rev or 0) if agg else 0
        budget_lab = float(agg.budget_lab or 0) if agg else 0
        budget_costs = float(agg.budget_costs or 0) if agg else 0
        contribution = budget_rev - budget_lab - budget_costs

        output.append(SiteAnnualSummaryRow(
            location_id=loc.id,
            location_name=loc.name or "",
            state=loc.state,
            total_prior_pet_days=prior_pd,
            total_budget_pet_days=budget_pd,
            total_prior_revenue=prior_rev,
            total_budget_revenue=budget_rev,
            total_budget_labour=budget_lab,
            total_budget_costs=budget_costs,
            budget_contribution=contribution,
            assumptions_status=a_status,
        ))

    return output


# ── PUT /budgets/{id}/sites/bulk-assumptions ──────────────────────────────────


@router.put("/{budget_id}/sites/bulk-assumptions")
async def bulk_update_site_assumptions(
    budget_id: uuid.UUID,
    payload: SiteBudgetAssumptionBulkUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_finance),
):
    """Apply the same growth rates to all sites."""
    from app.db.models.location import SiteBudgetAssumption

    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(SiteBudgetAssumption).where(
            SiteBudgetAssumption.version_id == budget_id,
            SiteBudgetAssumption.assumptions_locked.is_(False),
        )
    )
    assumptions = result.scalars().all()
    updated = 0

    for a in assumptions:
        if payload.price_growth_pct is not None:
            a.price_growth_pct = payload.price_growth_pct
        if payload.pet_day_growth_pct is not None:
            a.pet_day_growth_pct = payload.pet_day_growth_pct
        if payload.wage_increase_pct is not None:
            a.wage_increase_pct = payload.wage_increase_pct
        a.last_updated_by = user.id
        a.last_updated_at = datetime.now(timezone.utc)
        updated += 1

    await db.commit()
    return {"status": "ok", "updated": updated}


# ── GET /budgets/{id}/sites/{location_id} ────────────────────────────────────

SITE_LINE_ITEMS = [
    "boarding_revenue",
    "grooming_revenue",
    "other_revenue",
    "direct_wages",
    "direct_costs",
    "rent",
]

SITE_LINE_LABELS = {
    "boarding_revenue": "Boarding Revenue",
    "grooming_revenue": "Grooming Revenue",
    "other_revenue": "Other Revenue",
    "direct_wages": "Direct Wages",
    "direct_costs": "Direct Costs",
    "rent": "Rent",
}


@router.get("/{budget_id}/sites/{location_id}", response_model=SiteBudgetGridRead)
async def get_site_budget(
    budget_id: uuid.UUID,
    location_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    location = await db.get(Location, location_id)
    if location is None:
        raise HTTPException(status_code=404, detail="Location not found")

    # Load periods
    period_result = await db.execute(
        select(Period)
        .where(Period.fy_year == version.fy_year)
        .order_by(Period.fy_month)
    )
    periods = list(period_result.scalars().all())
    period_labels = [_period_label(p) for p in periods]
    month_label_map = {p.fy_month: _period_label(p) for p in periods}

    # Load weekly periods for this FY
    wp_result = await db.execute(
        select(WeeklyPeriod).where(WeeklyPeriod.fy_year == version.fy_year)
    )
    weekly_periods = list(wp_result.scalars().all())
    week_to_month: dict[uuid.UUID, int] = {}
    week_proration: dict[uuid.UUID, float] = {}
    for wp in weekly_periods:
        if wp.fy_month is not None and wp.days_this_week_in_fy_month is not None:
            week_to_month[wp.id] = wp.fy_month
            week_proration[wp.id] = wp.days_this_week_in_fy_month / 7.0

    # Load entries for this location
    entry_result = await db.execute(
        select(SiteBudgetEntry).where(
            SiteBudgetEntry.version_id == budget_id,
            SiteBudgetEntry.location_id == location_id,
        )
    )
    entries = list(entry_result.scalars().all())

    # Build line → month → amount (prorated from weekly)
    line_values: dict[str, dict[str, float]] = {
        li: {} for li in SITE_LINE_ITEMS
    }
    for entry in entries:
        li = entry.model_line_item
        if li not in line_values:
            continue
        if entry.week_id and entry.week_id in week_to_month:
            fy_month = week_to_month[entry.week_id]
            factor = week_proration.get(entry.week_id, 1.0)
            amount = float(entry.amount or 0) * factor
            label = month_label_map.get(fy_month, f"M{fy_month:02d}")
            line_values[li][label] = line_values[li].get(label, 0.0) + amount

    lines = [
        SiteBudgetLineRead(
            line_item=SITE_LINE_LABELS.get(li, li),
            values=line_values.get(li, {}),
        )
        for li in SITE_LINE_ITEMS
    ]

    return SiteBudgetGridRead(
        location_id=location_id,
        location_name=location.name or "",
        periods=period_labels,
        lines=lines,
    )


# ── PUT /budgets/{id}/sites/{location_id} ────────────────────────────────────


@router.put("/{budget_id}/sites/{location_id}", response_model=SiteBudgetGridRead)
async def save_site_budget(
    budget_id: uuid.UUID,
    location_id: uuid.UUID,
    payload: SiteBudgetSavePayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    location = await db.get(Location, location_id)
    if location is None:
        raise HTTPException(status_code=404, detail="Location not found")

    # Load periods + weekly periods
    period_result = await db.execute(
        select(Period)
        .where(Period.fy_year == version.fy_year)
        .order_by(Period.fy_month)
    )
    periods = list(period_result.scalars().all())
    label_to_month: dict[str, int] = {_period_label(p): p.fy_month for p in periods}

    wp_result = await db.execute(
        select(WeeklyPeriod)
        .where(WeeklyPeriod.fy_year == version.fy_year)
        .order_by(WeeklyPeriod.week_start_date)
    )
    weekly_periods = list(wp_result.scalars().all())

    # Group weekly periods by fy_month — pick the first week of each month
    # to store the monthly total (driver_type='manual')
    first_week_per_month: dict[int, WeeklyPeriod] = {}
    for wp in weekly_periods:
        if wp.fy_month is not None and wp.fy_month not in first_week_per_month:
            first_week_per_month[wp.fy_month] = wp

    now = datetime.now(timezone.utc)
    label_to_line = {v: k for k, v in SITE_LINE_LABELS.items()}

    # Delete existing entries for this version + location
    from sqlalchemy import delete
    await db.execute(
        delete(SiteBudgetEntry).where(
            SiteBudgetEntry.version_id == budget_id,
            SiteBudgetEntry.location_id == location_id,
        )
    )

    # Insert new entries — one per line_item per month
    for line in payload.lines:
        line_item = label_to_line.get(line.line_item, line.line_item)
        for label, amount in line.values.items():
            fy_month = label_to_month.get(label)
            if fy_month is None:
                continue
            week = first_week_per_month.get(fy_month)
            if week is None:
                continue
            db.add(SiteBudgetEntry(
                version_id=budget_id,
                location_id=location_id,
                model_line_item=line_item,
                week_id=week.id,
                amount=amount,
                driver_type="manual",
                entered_by=user.id,
                updated_at=now,
            ))

    await db.commit()

    # Trigger rollup in background
    from app.services.site_rollup_service import rollup_sites_to_entity
    try:
        await rollup_sites_to_entity(db, budget_id)
        await db.commit()
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Rollup failed after site save")

    return await get_site_budget(budget_id, location_id, db, user)


# ── GET /budgets/{id}/sites/import/template ───────────────────────────────────


@router.get("/{budget_id}/sites/import/template")
async def download_site_import_template(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(Location)
        .where(Location.is_active.is_(True))
        .order_by(Location.state, Location.name)
    )
    locations = list(result.scalars().all())

    period_result = await db.execute(
        select(Period)
        .where(Period.fy_year == version.fy_year)
        .order_by(Period.fy_month)
    )
    periods = list(period_result.scalars().all())
    period_labels = [_period_label(p) for p in periods]

    NAVY = "1F3D6E"
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    loc_font = Font(bold=True, size=10)
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Site Budgets FY{version.fy_year}"

    headers = ["Location", "Line Item"] + period_labels
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    for i in range(len(period_labels)):
        ws.column_dimensions[chr(67 + i)].width = 12

    row_idx = 2
    for loc in locations:
        loc_label = loc.name or loc.code or str(loc.id)
        for li_key in SITE_LINE_ITEMS:
            ws.cell(row=row_idx, column=1, value=loc_label).font = loc_font
            ws.cell(row=row_idx, column=2, value=SITE_LINE_LABELS[li_key])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    # Instructions sheet
    instr = wb.create_sheet(title="Instructions")
    instr.column_dimensions["A"].width = 80
    instructions = [
        f"Site Budget Import Template — FY{version.fy_year}",
        "",
        "Fill in the monthly amounts on the first sheet.",
        "- Location names must match exactly (or use location codes).",
        "- Line items: Boarding Revenue, Grooming Revenue, Other Revenue, Direct Wages, Direct Costs, Rent.",
        "- Leave cells blank or 0 for months with no budget.",
        "- You can delete rows for locations you don't need.",
        "- You can add rows — the import will auto-match location names.",
        "",
        "Then upload the file on the Site Budgets page using the Import button.",
    ]
    for i, text in enumerate(instructions, 1):
        instr.cell(row=i, column=1, value=text)
    instr.cell(row=1, column=1).font = Font(bold=True, size=12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kip_site_budget_template_FY{version.fy_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── POST /budgets/{id}/sites/import ──────────────────────────────────────────


def _normalise(s: str) -> str:
    """Lowercase, strip whitespace and common punctuation for fuzzy matching."""
    return s.lower().strip().replace("'", "").replace("'", "").replace("-", " ")


@router.post("/{budget_id}/sites/import")
async def import_site_budgets(
    budget_id: uuid.UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read file as Excel (.xlsx)")

    ws = wb.worksheets[0]
    rows_raw = list(ws.iter_rows(values_only=True))
    if len(rows_raw) < 2:
        raise HTTPException(status_code=400, detail="File has no data rows")

    header_row = rows_raw[0]
    data_rows = rows_raw[1:]

    # ── Map period columns ────────────────────────────────────────────────
    period_result = await db.execute(
        select(Period)
        .where(Period.fy_year == version.fy_year)
        .order_by(Period.fy_month)
    )
    periods = list(period_result.scalars().all())

    # Build multiple label variants for each month
    label_to_month: dict[str, int] = {}
    for p in periods:
        lbl = _period_label(p).lower()
        label_to_month[lbl] = p.fy_month
        # Also match "m01", "m1", "jul", "july", month number alone
        label_to_month[f"m{p.fy_month:02d}"] = p.fy_month
        label_to_month[f"m{p.fy_month}"] = p.fy_month
        month_name = MONTH_ABBR[p.fy_month - 1].lower()
        label_to_month[month_name] = p.fy_month

    col_to_month: dict[int, int] = {}
    for col_idx, hdr in enumerate(header_row):
        if hdr is None:
            continue
        normalised = str(hdr).strip().lower()
        # Try exact match first, then prefix match
        if normalised in label_to_month:
            col_to_month[col_idx] = label_to_month[normalised]
        else:
            for key, month in label_to_month.items():
                if normalised.startswith(key) or key.startswith(normalised):
                    col_to_month[col_idx] = month
                    break

    if not col_to_month:
        raise HTTPException(
            status_code=400,
            detail="Could not match any column headers to FY months. "
                   "Expected headers like Jul-25, Aug-25 or M01, M02, etc."
        )

    # ── Map locations ─────────────────────────────────────────────────────
    loc_result = await db.execute(select(Location))
    all_locations = list(loc_result.scalars().all())

    loc_by_name: dict[str, Location] = {}
    loc_by_code: dict[str, Location] = {}
    for loc in all_locations:
        if loc.name:
            loc_by_name[_normalise(loc.name)] = loc
        if loc.code:
            loc_by_code[_normalise(loc.code)] = loc

    # ── Map line items ────────────────────────────────────────────────────
    label_to_line: dict[str, str] = {}
    for key, label in SITE_LINE_LABELS.items():
        label_to_line[_normalise(label)] = key
    # Common alternative names
    label_to_line["boarding"] = "boarding_revenue"
    label_to_line["grooming"] = "grooming_revenue"
    label_to_line["other"] = "other_revenue"
    label_to_line["wages"] = "direct_wages"
    label_to_line["direct wage"] = "direct_wages"
    label_to_line["direct cost"] = "direct_costs"
    label_to_line["costs"] = "direct_costs"

    # ── Detect column layout ─────────────────────────────────────────────
    # Try to figure out which columns are "Location" and "Line Item"
    loc_col = None
    li_col = None
    for col_idx, hdr in enumerate(header_row):
        if hdr is None:
            continue
        h = str(hdr).strip().lower()
        if h in ("location", "site", "site name", "location name"):
            loc_col = col_idx
        elif h in ("line item", "line_item", "lineitem", "item", "category", "account"):
            li_col = col_idx

    # Default: first column is location, second is line item
    if loc_col is None:
        loc_col = 0
    if li_col is None:
        li_col = 1

    # ── Load weekly periods for insertion ─────────────────────────────────
    wp_result = await db.execute(
        select(WeeklyPeriod)
        .where(WeeklyPeriod.fy_year == version.fy_year)
        .order_by(WeeklyPeriod.week_start_date)
    )
    weekly_periods = list(wp_result.scalars().all())

    first_week_per_month: dict[int, WeeklyPeriod] = {}
    for wp in weekly_periods:
        if wp.fy_month is not None and wp.fy_month not in first_week_per_month:
            first_week_per_month[wp.fy_month] = wp

    now = datetime.now(timezone.utc)

    # ── Parse rows and insert ─────────────────────────────────────────────
    matched_locations: set[str] = set()
    unmatched_locations: set[str] = set()
    matched_line_items: set[str] = set()
    unmatched_line_items: set[str] = set()
    rows_imported = 0
    entries_created = 0
    locations_updated: set[uuid.UUID] = set()

    # Collect all data first, grouped by location
    import_data: dict[uuid.UUID, dict[str, dict[int, float]]] = defaultdict(
        lambda: defaultdict(dict)
    )

    for row in data_rows:
        if not row or all(v is None for v in row):
            continue

        loc_val = str(row[loc_col] or "").strip()
        li_val = str(row[li_col] or "").strip()

        if not loc_val:
            continue

        # Match location
        norm_loc = _normalise(loc_val)
        location = loc_by_name.get(norm_loc) or loc_by_code.get(norm_loc)
        if not location:
            # Partial match: check if any known name contains this value or vice versa
            for known_name, known_loc in loc_by_name.items():
                if norm_loc in known_name or known_name in norm_loc:
                    location = known_loc
                    break
            if not location:
                for known_code, known_loc in loc_by_code.items():
                    if norm_loc == known_code or known_code.startswith(norm_loc):
                        location = known_loc
                        break

        if not location:
            unmatched_locations.add(loc_val)
            continue

        matched_locations.add(loc_val)

        # Match line item
        norm_li = _normalise(li_val)
        line_key = label_to_line.get(norm_li)
        if not line_key:
            # Partial match
            for known_li, known_key in label_to_line.items():
                if norm_li in known_li or known_li in norm_li:
                    line_key = known_key
                    break

        if not line_key:
            unmatched_line_items.add(li_val)
            continue

        matched_line_items.add(li_val)

        # Extract amounts
        for col_idx, fy_month in col_to_month.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    amount = float(row[col_idx])
                except (ValueError, TypeError):
                    continue
                import_data[location.id][line_key][fy_month] = amount

        rows_imported += 1

    # ── Delete existing and insert ────────────────────────────────────────
    from sqlalchemy import delete

    for loc_id, line_items in import_data.items():
        # Delete existing entries for this version + location
        await db.execute(
            delete(SiteBudgetEntry).where(
                SiteBudgetEntry.version_id == budget_id,
                SiteBudgetEntry.location_id == loc_id,
            )
        )
        locations_updated.add(loc_id)

        for line_key, month_amounts in line_items.items():
            for fy_month, amount in month_amounts.items():
                week = first_week_per_month.get(fy_month)
                if week is None:
                    continue
                db.add(SiteBudgetEntry(
                    version_id=budget_id,
                    location_id=loc_id,
                    model_line_item=line_key,
                    week_id=week.id,
                    amount=amount,
                    driver_type="manual",
                    entered_by=user.id,
                    updated_at=now,
                ))
                entries_created += 1

    await db.commit()

    # Run rollup
    from app.services.site_rollup_service import rollup_sites_to_entity
    try:
        await rollup_sites_to_entity(db, budget_id)
        await db.commit()
    except Exception:
        import logging
        logging.getLogger(__name__).exception("Rollup failed after bulk import")

    return {
        "status": "success",
        "rows_imported": rows_imported,
        "entries_created": entries_created,
        "locations_updated": len(locations_updated),
        "matched_locations": sorted(matched_locations),
        "unmatched_locations": sorted(unmatched_locations),
        "matched_line_items": sorted(matched_line_items),
        "unmatched_line_items": sorted(unmatched_line_items),
        "periods_matched": len(col_to_month),
    }


# ── GET /budgets/{id}/sites/rollup ───────────────────────────────────────────


@router.get("/{budget_id}/site-rollup", response_model=list[SiteRollupRow])
async def get_site_rollup(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    # Load periods
    period_result = await db.execute(
        select(Period)
        .where(Period.fy_year == version.fy_year)
        .order_by(Period.fy_month)
    )
    periods = list(period_result.scalars().all())
    month_label: dict[int, str] = {p.fy_month: _period_label(p) for p in periods}

    # Load weekly periods for proration
    wp_result = await db.execute(
        select(WeeklyPeriod).where(WeeklyPeriod.fy_year == version.fy_year)
    )
    weekly_periods = list(wp_result.scalars().all())
    week_to_month: dict[uuid.UUID, int] = {}
    week_proration: dict[uuid.UUID, float] = {}
    for wp in weekly_periods:
        if wp.fy_month is not None and wp.days_this_week_in_fy_month is not None:
            week_to_month[wp.id] = wp.fy_month
            week_proration[wp.id] = wp.days_this_week_in_fy_month / 7.0

    # Load locations → entity mapping
    loc_result = await db.execute(select(Location).where(Location.is_active.is_(True)))
    locations = list(loc_result.scalars().all())
    loc_entity: dict[uuid.UUID, uuid.UUID] = {
        loc.id: loc.entity_id for loc in locations if loc.entity_id
    }

    # Load entities
    ent_result = await db.execute(
        select(Entity).where(Entity.is_active.is_(True))
    )
    entity_map = {e.id: e for e in ent_result.scalars().all()}

    # Load site entries
    entry_result = await db.execute(
        select(SiteBudgetEntry).where(SiteBudgetEntry.version_id == budget_id)
    )
    entries = list(entry_result.scalars().all())

    # Aggregate by entity × line_item × month
    agg: dict[tuple[uuid.UUID, str], dict[str, float]] = defaultdict(
        lambda: defaultdict(float)
    )
    for entry in entries:
        entity_id = loc_entity.get(entry.location_id)
        if entity_id is None:
            continue
        if entry.week_id and entry.week_id in week_to_month:
            fy_month = week_to_month[entry.week_id]
            factor = week_proration.get(entry.week_id, 1.0)
            amount = float(entry.amount or 0) * factor
        else:
            continue
        label = month_label.get(fy_month, f"M{fy_month:02d}")
        agg[(entity_id, entry.model_line_item or "")][label] += amount

    # Load model assumptions for comparison
    ma_result = await db.execute(
        select(ModelAssumption).where(
            ModelAssumption.budget_version_id == budget_id
        )
    )
    model_assumptions = list(ma_result.scalars().all())
    ma_map: dict[tuple[uuid.UUID | None, str], dict] = {}
    for ma in model_assumptions:
        ma_map[(ma.entity_id, ma.assumption_key)] = ma.assumption_value

    rows: list[SiteRollupRow] = []
    for (entity_id, line_item), site_totals in sorted(
        agg.items(), key=lambda x: (str(x[0][0]), x[0][1])
    ):
        entity = entity_map.get(entity_id)
        if not entity:
            continue

        assumption_key = f"site_rollup.{line_item}"
        model_vals = ma_map.get((entity_id, assumption_key), {})

        variance: dict[str, float] = {}
        for lbl in site_totals:
            sv = site_totals.get(lbl, 0.0)
            mv = float(model_vals.get(lbl, 0.0)) if isinstance(model_vals, dict) else 0.0
            variance[lbl] = sv - mv

        rows.append(SiteRollupRow(
            entity_code=entity.code or "",
            entity_name=entity.name,
            line_item=SITE_LINE_LABELS.get(line_item, line_item),
            site_total=dict(site_totals),
            model_assumption={
                k: float(v)
                for k, v in model_vals.items()
                if isinstance(v, (int, float))
            } if isinstance(model_vals, dict) else {},
            variance=variance,
        ))

    return rows


# ── POST /budgets/{id}/calculate ─────────────────────────────────────────────


@router.post(
    "/{budget_id}/calculate",
    response_model=CalculationTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def trigger_calculation(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    from app.worker import run_model_task

    task_id = f"model-{budget_id}"
    run_model_task.apply_async(args=[str(budget_id)], task_id=task_id)
    return CalculationTriggerResponse(task_id=task_id, status="queued")


# ── GET /budgets/{id}/status ─────────────────────────────────────────────────


@router.get("/{budget_id}/status", response_model=CalculationStatusResponse)
async def get_calculation_status(
    budget_id: uuid.UUID,
    _user: User = Depends(require_finance),
):
    from app.worker import celery_app

    # Find the most recent task for this budget version.
    # In a production system you'd track task_id on the version row;
    # for now inspect via Celery backend.
    task_key = f"model-{budget_id}"
    result = AsyncResult(task_key, app=celery_app)

    if result.state == "PENDING":
        return CalculationStatusResponse(
            task_id=task_key, status="no_task_found"
        )

    return CalculationStatusResponse(
        task_id=task_key,
        status=result.state.lower(),
        result=result.result if result.successful() else None,
    )


# ── GET /budgets/{id}/output/is ──────────────────────────────────────────────


@router.get("/{budget_id}/output/is", response_model=ModelOutputResponse)
async def get_budget_is(
    budget_id: uuid.UUID,
    fy_month: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    return await _get_model_output(db, budget_id, Statement.is_, fy_month)


@router.get("/{budget_id}/output/bs", response_model=ModelOutputResponse)
async def get_budget_bs(
    budget_id: uuid.UUID,
    fy_month: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    return await _get_model_output(db, budget_id, Statement.bs, fy_month)


@router.get("/{budget_id}/output/cf", response_model=ModelOutputResponse)
async def get_budget_cf(
    budget_id: uuid.UUID,
    fy_month: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    return await _get_model_output(db, budget_id, Statement.cf, fy_month)


# ── Shared output renderer ──────────────────────────────────────────────────


async def _get_model_output(
    db: AsyncSession,
    budget_id: uuid.UUID,
    statement: Statement,
    fy_month: int | None,
) -> ModelOutputResponse:
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    fy_year = version.fy_year

    # Resolve periods
    if fy_month is not None:
        result = await db.execute(
            select(Period)
            .where(Period.fy_year == fy_year, Period.fy_month <= fy_month)
            .order_by(Period.fy_month)
        )
        all_periods = list(result.scalars().all())
        target = next((p for p in all_periods if p.fy_month == fy_month), None)
        if target is None:
            raise HTTPException(status_code=404, detail="Period not found")
        period_ids = [p.id for p in all_periods]
        target_label = _period_label(target)
        column_labels = [target_label, "YTD"]
        period_id_to_label = {p.id: _period_label(p) for p in all_periods}
    else:
        result = await db.execute(
            select(Period)
            .where(Period.fy_year == fy_year)
            .order_by(Period.fy_month)
        )
        all_periods = list(result.scalars().all())
        if not all_periods:
            raise HTTPException(status_code=404, detail="No periods for this FY")
        period_ids = [p.id for p in all_periods]
        column_labels = [_period_label(p) for p in all_periods]
        period_id_to_label = {p.id: _period_label(p) for p in all_periods}

    # Load accounts for the statement (or CF pseudo-accounts)
    is_cf = statement == Statement.cf
    if is_cf:
        cf_codes = ["CF-OPERATING", "CF-INVESTING", "CF-FINANCING", "CF-NET"]
        result = await db.execute(
            select(Account).where(Account.code.in_(cf_codes))
        )
        accounts = list(result.scalars().all())
        if not accounts:
            accounts = _build_cf_pseudo_accounts(cf_codes)
    else:
        result = await db.execute(
            select(Account)
            .where(Account.statement == statement)
            .order_by(Account.sort_order)
        )
        accounts = list(result.scalars().all())

    account_ids = [a.id for a in accounts]

    # Load model outputs
    result = await db.execute(
        select(ModelOutput).where(
            ModelOutput.version_id == budget_id,
            ModelOutput.period_id.in_(period_ids),
            ModelOutput.account_id.in_(account_ids),
        )
    )
    model_outputs = result.scalars().all()

    # Load entities
    result = await db.execute(select(Entity))
    entities_map = {e.id: e for e in result.scalars().all()}

    # Index outputs
    group_totals: dict[uuid.UUID, dict[str, float]] = defaultdict(
        lambda: defaultdict(float)
    )
    entity_amounts: dict[uuid.UUID, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(float))
    )

    for mo in model_outputs:
        label = period_id_to_label.get(mo.period_id)
        if label is None:
            continue
        if mo.entity_id is None:
            group_totals[mo.account_id][label] += float(mo.amount)
        else:
            ent = entities_map.get(mo.entity_id)
            ecode = ent.code if ent else "?"
            entity_amounts[mo.account_id][ecode][label] += float(mo.amount)

    # Build rows with section headers
    rows: list[ModelOutputRow] = []

    if is_cf:
        for section, codes in CF_SECTIONS.items():
            rows.append(ModelOutputRow(
                account_code="",
                label=section,
                is_section_header=True,
            ))
            for acct in accounts:
                if acct.code not in codes:
                    continue
                vals = _build_values(
                    acct, group_totals, column_labels, fy_month, period_id_to_label
                )
                eb = _build_entity_breakdown(
                    acct, entity_amounts, column_labels, fy_month, period_id_to_label
                )
                rows.append(ModelOutputRow(
                    account_code=acct.code,
                    label=acct.name,
                    is_subtotal=getattr(acct, "is_subtotal", False),
                    indent_level=0 if getattr(acct, "is_subtotal", False) else 1,
                    values=vals,
                    entity_breakdown=eb,
                ))
    else:
        sections = BS_SECTIONS if statement == Statement.bs else IS_SECTIONS
        type_to_section: dict[AccountType, str] = {}
        for sec, types in sections.items():
            for t in types:
                type_to_section[t] = sec

        seen_sections: set[str] = set()
        for acct in accounts:
            sec = type_to_section.get(acct.account_type)
            if sec and sec not in seen_sections:
                seen_sections.add(sec)
                rows.append(ModelOutputRow(
                    account_code="",
                    label=sec,
                    is_section_header=True,
                ))

            vals = _build_values(
                acct, group_totals, column_labels, fy_month, period_id_to_label
            )
            eb = _build_entity_breakdown(
                acct, entity_amounts, column_labels, fy_month, period_id_to_label
            )
            rows.append(ModelOutputRow(
                account_code=acct.code,
                label=acct.name,
                is_subtotal=acct.is_subtotal,
                indent_level=0 if acct.is_subtotal else 1,
                values=vals,
                entity_breakdown=eb,
            ))

    return ModelOutputResponse(
        version_id=budget_id,
        fy_year=fy_year,
        statement=statement.value,
        periods=column_labels,
        rows=rows,
    )


def _build_values(
    acct: Account,
    group_totals: dict,
    column_labels: list[str],
    fy_month: int | None,
    period_id_to_label: dict,
) -> dict[str, float]:
    month_totals = group_totals.get(acct.id, {})
    if fy_month is not None:
        target_label = column_labels[0]
        month_val = month_totals.get(target_label, 0.0)
        ytd_sum = sum(month_totals.values())
        return {target_label: month_val, "YTD": ytd_sum}
    return {lbl: month_totals.get(lbl, 0.0) for lbl in column_labels}


def _build_entity_breakdown(
    acct: Account,
    entity_amounts: dict,
    column_labels: list[str],
    fy_month: int | None,
    period_id_to_label: dict,
) -> dict[str, dict[str, float]]:
    eb: dict[str, dict[str, float]] = {}
    for ecode, period_map in entity_amounts.get(acct.id, {}).items():
        if fy_month is not None:
            target_label = column_labels[0]
            eb[ecode] = {
                target_label: period_map.get(target_label, 0.0),
                "YTD": sum(period_map.values()),
            }
        else:
            eb[ecode] = {lbl: period_map.get(lbl, 0.0) for lbl in column_labels}
    return eb


def _build_cf_pseudo_accounts(codes: list[str]) -> list:
    """If CF accounts don't exist in the DB yet, create lightweight objects."""
    from types import SimpleNamespace

    labels = {
        "CF-OPERATING": "Operating Cash Flow",
        "CF-INVESTING": "Investing Cash Flow",
        "CF-FINANCING": "Financing Cash Flow",
        "CF-NET": "Net Cash Flow",
    }
    return [
        SimpleNamespace(
            id=uuid.uuid5(uuid.NAMESPACE_DNS, c),
            code=c,
            name=labels.get(c, c),
            is_subtotal=(c == "CF-NET"),
            account_type=None,
        )
        for c in codes
    ]


# ══════════════════════════════════════════════════════════════════════════════
# Site Budget Engine Endpoints (Operational Budget)
# ══════════════════════════════════════════════════════════════════════════════


@router.get(
    "/{budget_id}/sites/{location_id}/assumptions",
    response_model=SiteBudgetAssumptionRead,
)
async def get_site_assumptions(
    budget_id: uuid.UUID,
    location_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    """Get site budget assumptions, auto-populating from prior year if needed."""
    from sqlalchemy import func as sqla_func

    from app.db.models.location import SiteBudgetAssumption
    from app.db.models.pet_days import SitePetDay

    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(SiteBudgetAssumption).where(
            SiteBudgetAssumption.version_id == budget_id,
            SiteBudgetAssumption.location_id == location_id,
        )
    )
    assumptions = result.scalar_one_or_none()

    if assumptions is None:
        from app.services.site_budget_engine import _auto_populate_assumptions
        assumptions = await _auto_populate_assumptions(
            db, budget_id, location_id, version.fy_year, version.fy_year - 1,
        )
        await db.commit()

    prior_fy = version.fy_year - 1
    wp_result = await db.execute(
        select(WeeklyPeriod).where(WeeklyPeriod.fy_year == prior_fy)
    )
    prior_weeks = wp_result.scalars().all()
    prior_avg_price = 0.0
    prior_total_pd = 0
    prior_avg_wage = 35.0

    if prior_weeks:
        date_min = min(w.week_start_date for w in prior_weeks)
        date_max = max(w.week_end_date or w.week_start_date for w in prior_weeks)

        pd_result = await db.execute(
            select(
                sqla_func.sum(SitePetDay.pet_days),
                sqla_func.sum(SitePetDay.revenue_aud),
            ).where(
                SitePetDay.location_id == location_id,
                SitePetDay.date >= date_min,
                SitePetDay.date <= date_max,
            )
        )
        row = pd_result.one()
        total_pd, total_rev = row
        prior_total_pd = int(total_pd or 0)
        if prior_total_pd > 0 and total_rev:
            prior_avg_price = float(total_rev) / prior_total_pd

        from app.services.site_budget_engine import _get_prior_year_avg_wage
        prior_avg_wage = float(await _get_prior_year_avg_wage(db, location_id, prior_fy))

    return SiteBudgetAssumptionRead(
        id=assumptions.id,
        version_id=assumptions.version_id,
        location_id=assumptions.location_id,
        fy_year=assumptions.fy_year,
        price_growth_pct=float(assumptions.price_growth_pct) if assumptions.price_growth_pct else None,
        pet_day_growth_pct=float(assumptions.pet_day_growth_pct) if assumptions.pet_day_growth_pct else None,
        bath_price=float(assumptions.bath_price) if assumptions.bath_price else None,
        other_services_per_pet_day=float(assumptions.other_services_per_pet_day) if assumptions.other_services_per_pet_day else None,
        membership_pct_revenue=float(assumptions.membership_pct_revenue) if assumptions.membership_pct_revenue else None,
        mpp_mins=float(assumptions.mpp_mins) if assumptions.mpp_mins else None,
        min_daily_hours=float(assumptions.min_daily_hours) if assumptions.min_daily_hours else None,
        wage_increase_pct=float(assumptions.wage_increase_pct) if assumptions.wage_increase_pct else None,
        cogs_pct=float(assumptions.cogs_pct) if assumptions.cogs_pct else None,
        rent_monthly=float(assumptions.rent_monthly) if assumptions.rent_monthly else None,
        rent_growth_pct=float(assumptions.rent_growth_pct) if assumptions.rent_growth_pct else None,
        utilities_monthly=float(assumptions.utilities_monthly) if assumptions.utilities_monthly else None,
        utilities_growth_pct=float(assumptions.utilities_growth_pct) if assumptions.utilities_growth_pct else None,
        rm_monthly=float(assumptions.rm_monthly) if assumptions.rm_monthly else None,
        rm_growth_pct=float(assumptions.rm_growth_pct) if assumptions.rm_growth_pct else None,
        it_monthly=float(assumptions.it_monthly) if assumptions.it_monthly else None,
        it_growth_pct=float(assumptions.it_growth_pct) if assumptions.it_growth_pct else None,
        general_monthly=float(assumptions.general_monthly) if assumptions.general_monthly else None,
        general_growth_pct=float(assumptions.general_growth_pct) if assumptions.general_growth_pct else None,
        advertising_pct_revenue=float(assumptions.advertising_pct_revenue) if assumptions.advertising_pct_revenue else None,
        assumptions_locked=assumptions.assumptions_locked,
        prior_year_avg_price=prior_avg_price,
        prior_year_total_pet_days=prior_total_pd,
        prior_year_avg_wage=prior_avg_wage,
    )


@router.put(
    "/{budget_id}/sites/{location_id}/assumptions",
    response_model=SiteBudgetAssumptionRead,
)
async def save_site_assumptions(
    budget_id: uuid.UUID,
    location_id: uuid.UUID,
    payload: SiteBudgetAssumptionUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_finance),
):
    """Save site budget assumptions."""
    from app.db.models.location import SiteBudgetAssumption

    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(SiteBudgetAssumption).where(
            SiteBudgetAssumption.version_id == budget_id,
            SiteBudgetAssumption.location_id == location_id,
        )
    )
    assumptions = result.scalar_one_or_none()

    if assumptions is None:
        assumptions = SiteBudgetAssumption(
            version_id=budget_id,
            location_id=location_id,
            fy_year=version.fy_year,
        )
        db.add(assumptions)

    update_data = payload.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if value is not None:
            setattr(assumptions, field, value)

    assumptions.last_updated_by = user.id
    assumptions.last_updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(assumptions)

    return await get_site_assumptions(budget_id, location_id, db, user)


@router.post(
    "/{budget_id}/sites/{location_id}/calculate",
    response_model=CalculationTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def trigger_site_calculation(
    budget_id: uuid.UUID,
    location_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    """Calculate weekly budget for a single site."""
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    from app.services.site_budget_engine import calculate_site_weekly_budget

    result = await calculate_site_weekly_budget(budget_id, location_id, db)
    await db.commit()

    return CalculationTriggerResponse(
        task_id=f"site-{budget_id}-{location_id}",
        status="completed",
    )


@router.post(
    "/{budget_id}/calculate-all-sites",
    response_model=CalculationTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def trigger_all_sites_calculation(
    budget_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    """Calculate weekly budget for all active sites, then roll up and run model."""
    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    from app.worker import calculate_all_sites_task

    task_id = f"all-sites-{budget_id}"
    calculate_all_sites_task.apply_async(args=[str(budget_id)], task_id=task_id)
    return CalculationTriggerResponse(task_id=task_id, status="queued")


@router.get(
    "/{budget_id}/sites/{location_id}/weekly",
    response_model=list[SiteWeeklyBudgetRow],
)
async def get_site_weekly_budget(
    budget_id: uuid.UUID,
    location_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    """Weekly budget grid for a single site."""
    from app.db.models.location import SiteWeeklyBudget

    version = await db.get(BudgetVersion, budget_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    result = await db.execute(
        select(WeeklyPeriod)
        .where(WeeklyPeriod.fy_year == version.fy_year)
        .order_by(WeeklyPeriod.week_start_date)
    )
    weeks = result.scalars().all()
    week_map = {w.id: w for w in weeks}

    result = await db.execute(
        select(SiteWeeklyBudget).where(
            SiteWeeklyBudget.version_id == budget_id,
            SiteWeeklyBudget.location_id == location_id,
        )
    )
    budget_rows = {r.week_id: r for r in result.scalars().all()}

    output: list[SiteWeeklyBudgetRow] = []
    month_subtotals: dict[int, dict] = defaultdict(lambda: {
        "prior_revenue": 0.0, "budget_revenue": 0.0,
        "budget_labour": 0.0, "budget_cogs": 0.0,
        "budget_rent": 0.0, "prior_boarding": 0, "prior_daycare": 0,
        "budget_boarding": 0, "budget_daycare": 0,
    })
    last_month = None

    for w in weeks:
        br = budget_rows.get(w.id)
        current_month = w.fy_month

        if last_month is not None and current_month != last_month and last_month in month_subtotals:
            sub = month_subtotals[last_month]
            contribution = sub["budget_revenue"] - sub["budget_labour"] - sub["budget_cogs"] - sub["budget_rent"]
            output.append(SiteWeeklyBudgetRow(
                week_id=uuid.UUID(int=0),
                week_label=f"Month {last_month} Total",
                fy_month=last_month,
                prior_year_revenue=sub["prior_revenue"],
                budget_revenue=sub["budget_revenue"],
                budget_labour=sub["budget_labour"],
                is_month_subtotal=True,
            ))

        row = SiteWeeklyBudgetRow(
            week_id=w.id,
            week_label=w.week_label,
            week_start=str(w.week_start_date) if w.week_start_date else None,
            week_end=str(w.week_end_date) if w.week_end_date else None,
            fy_month=w.fy_month,
        )

        if br:
            row.prior_year_pet_days_boarding = br.prior_year_pet_days_boarding
            row.prior_year_pet_days_daycare = br.prior_year_pet_days_daycare
            row.prior_year_pet_days_grooming = br.prior_year_pet_days_grooming
            row.prior_year_pet_days_wash = br.prior_year_pet_days_wash
            row.prior_year_pet_days_training = br.prior_year_pet_days_training
            row.prior_year_revenue = float(br.prior_year_revenue or 0)
            row.budget_pet_days_boarding = br.budget_pet_days_boarding
            row.budget_pet_days_daycare = br.budget_pet_days_daycare
            row.budget_pet_days_grooming = br.budget_pet_days_grooming
            row.budget_pet_days_wash = br.budget_pet_days_wash
            row.budget_pet_days_training = br.budget_pet_days_training
            row.budget_revenue = float(br.override_revenue if br.is_overridden and br.override_revenue else br.budget_revenue or 0)
            row.budget_labour = float(br.override_labour if br.is_overridden and br.override_labour else br.budget_labour or 0)
            row.budget_cogs = float(br.budget_cogs or 0)
            row.budget_rent = float(br.budget_rent or 0)
            row.budget_utilities = float(br.budget_utilities or 0)
            row.budget_rm = float(br.budget_rm or 0)
            row.budget_it = float(br.budget_it or 0)
            row.budget_general = float(br.budget_general or 0)
            row.budget_advertising = float(br.budget_advertising or 0)
            row.is_overridden = br.is_overridden
            row.override_revenue = float(br.override_revenue) if br.override_revenue else None
            row.override_labour = float(br.override_labour) if br.override_labour else None

            if current_month is not None:
                sub = month_subtotals[current_month]
                sub["prior_revenue"] += float(br.prior_year_revenue or 0)
                sub["budget_revenue"] += row.budget_revenue
                sub["budget_labour"] += row.budget_labour
                sub["budget_cogs"] += float(br.budget_cogs or 0)
                sub["budget_rent"] += float(br.budget_rent or 0)

        output.append(row)
        last_month = current_month

    if last_month is not None and last_month in month_subtotals:
        sub = month_subtotals[last_month]
        output.append(SiteWeeklyBudgetRow(
            week_id=uuid.UUID(int=0),
            week_label=f"Month {last_month} Total",
            fy_month=last_month,
            prior_year_revenue=sub["prior_revenue"],
            budget_revenue=sub["budget_revenue"],
            budget_labour=sub["budget_labour"],
            is_month_subtotal=True,
        ))

    return output


@router.put(
    "/{budget_id}/sites/{location_id}/weekly/{week_id}/override",
)
async def override_site_weekly_budget(
    budget_id: uuid.UUID,
    location_id: uuid.UUID,
    week_id: uuid.UUID,
    payload: SiteWeeklyOverridePayload,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    """Override revenue/labour for a single week."""
    from app.db.models.location import SiteWeeklyBudget

    result = await db.execute(
        select(SiteWeeklyBudget).where(
            SiteWeeklyBudget.version_id == budget_id,
            SiteWeeklyBudget.location_id == location_id,
            SiteWeeklyBudget.week_id == week_id,
        )
    )
    row = result.scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="Weekly budget row not found")

    row.is_overridden = payload.is_overridden
    row.override_revenue = payload.override_revenue
    row.override_labour = payload.override_labour
    row.calculated_at = datetime.now(timezone.utc)
    await db.commit()

    return {"status": "ok", "week_id": str(week_id), "is_overridden": row.is_overridden}
