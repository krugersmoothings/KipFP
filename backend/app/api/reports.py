"""Variance reporting, Excel/PDF export, and management pack generation."""

from __future__ import annotations

import io
import uuid
from collections import defaultdict
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_db, require_finance
from app.db.models.account import Account, AccountType, NormalBalance, Statement
from app.db.models.budget import (
    BudgetVersion,
    ModelAssumption,
    ModelOutput,
    ReportCommentary,
    VersionStatus,
    VersionType,
)
from app.db.models.consolidation import ConsolidatedActual
from app.db.models.entity import Entity
from app.db.models.period import Period
from app.db.models.user import User
from app.services.aasb16_helpers import compute_aasb16_by_account_period
from app.schemas.reports import (
    CommentaryPayload,
    CommentaryRead,
    ExportRequest,
    ManagementPackRequest,
    VarianceReportResponse,
    VarianceRow,
)

router = APIRouter(prefix="/reports", tags=["reports"])

MONTH_ABBR = [
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
]

IS_SECTIONS: dict[str, list[AccountType]] = {
    "Revenue": [AccountType.income, AccountType.cogs],
    "Operating Expenses": [AccountType.opex],
    "Depreciation & Amortisation": [AccountType.depreciation],
    "Interest": [AccountType.interest],
    "Tax": [AccountType.tax],
}

EXPENSE_TYPES = {AccountType.cogs, AccountType.opex, AccountType.depreciation, AccountType.tax}


def _period_label(period: Period) -> str:
    if period.period_start:
        return period.period_start.strftime("%b-%y")
    cal_year = period.fy_year - 1 if period.fy_month <= 6 else period.fy_year
    return f"{MONTH_ABBR[period.fy_month - 1]}-{cal_year % 100:02d}"


def _is_expense(acct: Account) -> bool:
    return acct.account_type in EXPENSE_TYPES


def _compute_variance(actual: float, budget: float, is_expense: bool) -> tuple[float, float | None, bool | None]:
    var_abs = actual - budget
    var_pct = (var_abs / budget * 100) if budget != 0 else None
    if budget == 0 and actual == 0:
        is_fav = None
    else:
        # Credit-normal convention: lower value is favourable for both
        # income (more negative = higher revenue) and expenses (lower cost).
        is_fav = actual < budget
    return var_abs, var_pct, is_fav


# ── GET /reports/variance ────────────────────────────────────────────────────


@router.get("/variance", response_model=VarianceReportResponse)
async def get_variance_report(
    fy_year: int = Query(...),
    fy_month: int | None = Query(None, description="Omit or pass 0 for YTD; -1 for full year"),
    version_id: uuid.UUID = Query(...),
    include_aasb16: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    version = await db.get(BudgetVersion, version_id)
    if version is None:
        raise HTTPException(status_code=404, detail="Budget version not found")

    # Determine view mode and periods
    if fy_month is not None and fy_month > 0:
        view_mode = "monthly"
        result = await db.execute(
            select(Period)
            .where(Period.fy_year == fy_year, Period.fy_month == fy_month)
        )
        target_period = result.scalar_one_or_none()
        if target_period is None:
            raise HTTPException(status_code=404, detail="Period not found")
        period_ids = [target_period.id]
        period_label = _period_label(target_period)
    elif fy_month is not None and fy_month == -1:
        view_mode = "full_year"
        result = await db.execute(
            select(Period)
            .where(Period.fy_year == fy_year, Period.fy_month >= 1)
            .order_by(Period.fy_month)
        )
        all_periods = list(result.scalars().all())
        period_ids = [p.id for p in all_periods]
        period_label = f"FY{fy_year} Full Year"
    else:
        view_mode = "ytd"
        latest_result = await db.execute(
            select(func.max(Period.fy_month))
            .join(ConsolidatedActual, ConsolidatedActual.period_id == Period.id)
            .where(
                Period.fy_year == fy_year,
                Period.fy_month >= 1,
                ConsolidatedActual.is_group_total.is_(True),
            )
        )
        ytd_cutoff = latest_result.scalar() or 12
        result = await db.execute(
            select(Period)
            .where(Period.fy_year == fy_year, Period.fy_month >= 1, Period.fy_month <= ytd_cutoff)
            .order_by(Period.fy_month)
        )
        all_periods = list(result.scalars().all())
        period_ids = [p.id for p in all_periods]
        period_label = f"FY{fy_year} YTD M{ytd_cutoff:02d}"

    # Load prior year periods for PCP comparison
    prior_fy = fy_year - 1
    if view_mode == "monthly" and fy_month:
        pcp_result = await db.execute(
            select(Period).where(Period.fy_year == prior_fy, Period.fy_month == fy_month)
        )
        pcp_periods = list(pcp_result.scalars().all())
    elif view_mode == "ytd":
        pcp_result = await db.execute(
            select(Period).where(
                Period.fy_year == prior_fy, Period.fy_month >= 1, Period.fy_month <= ytd_cutoff,
            ).order_by(Period.fy_month)
        )
        pcp_periods = list(pcp_result.scalars().all())
    else:
        pcp_result = await db.execute(
            select(Period).where(Period.fy_year == prior_fy).order_by(Period.fy_month)
        )
        pcp_periods = list(pcp_result.scalars().all())
    pcp_period_ids = [p.id for p in pcp_periods]

    # Load IS accounts
    result = await db.execute(
        select(Account)
        .where(Account.statement == Statement.is_)
        .order_by(Account.sort_order)
    )
    accounts = list(result.scalars().all())
    account_ids = [a.id for a in accounts]

    # Load actuals (current year)
    actual_totals: dict[uuid.UUID, float] = defaultdict(float)
    if period_ids:
        result = await db.execute(
            select(ConsolidatedActual).where(
                ConsolidatedActual.period_id.in_(period_ids),
                ConsolidatedActual.is_group_total.is_(True),
                ConsolidatedActual.account_id.in_(account_ids),
            )
        )
        for act in result.scalars().all():
            actual_totals[act.account_id] += float(act.amount)

    # AASB16 adjustment on actuals
    if not include_aasb16 and period_ids:
        aasb16_adj = await compute_aasb16_by_account_period(db, period_ids)
        for acct_id in account_ids:
            if acct_id in aasb16_adj:
                for pid in period_ids:
                    actual_totals[acct_id] -= aasb16_adj[acct_id].get(pid, 0.0)

    # Load budget outputs
    budget_totals: dict[uuid.UUID, float] = defaultdict(float)
    if period_ids:
        result = await db.execute(
            select(ModelOutput).where(
                ModelOutput.version_id == version_id,
                ModelOutput.period_id.in_(period_ids),
                ModelOutput.account_id.in_(account_ids),
                ModelOutput.entity_id.is_(None),
            )
        )
        for mo in result.scalars().all():
            budget_totals[mo.account_id] += float(mo.amount)

    # Load prior year actuals
    pcp_totals: dict[uuid.UUID, float] = defaultdict(float)
    if pcp_period_ids:
        result = await db.execute(
            select(ConsolidatedActual).where(
                ConsolidatedActual.period_id.in_(pcp_period_ids),
                ConsolidatedActual.is_group_total.is_(True),
                ConsolidatedActual.account_id.in_(account_ids),
            )
        )
        for act in result.scalars().all():
            pcp_totals[act.account_id] += float(act.amount)

    # Load commentary (keyed by version + account, period_id always NULL)
    commentary_map: dict[uuid.UUID, str] = {}
    result = await db.execute(
        select(ReportCommentary).where(
            ReportCommentary.version_id == version_id,
            ReportCommentary.period_id.is_(None),
        )
    )
    for c in result.scalars().all():
        commentary_map[c.account_id] = c.comment or ""

    # Build rows with section headers
    type_to_section: dict[AccountType, str] = {}
    for sec, types in IS_SECTIONS.items():
        for t in types:
            type_to_section[t] = sec

    rows: list[VarianceRow] = []
    seen_sections: set[str] = set()

    for acct in accounts:
        sec = type_to_section.get(acct.account_type)
        if sec and sec not in seen_sections:
            seen_sections.add(sec)
            rows.append(VarianceRow(
                account_code="",
                label=sec,
                is_section_header=True,
            ))

        actual = actual_totals.get(acct.id, 0.0)
        budget = budget_totals.get(acct.id, 0.0)
        pcp = pcp_totals.get(acct.id, 0.0)
        expense = _is_expense(acct)

        var_abs, var_pct, is_fav = _compute_variance(actual, budget, expense)
        pcp_abs = actual - pcp
        pcp_pct = (pcp_abs / pcp * 100) if pcp != 0 else None

        rows.append(VarianceRow(
            account_id=acct.id,
            account_code=acct.code,
            label=acct.name,
            is_subtotal=acct.is_subtotal,
            indent_level=0 if acct.is_subtotal else 1,
            actual=actual,
            budget=budget,
            variance_abs=var_abs,
            variance_pct=var_pct,
            is_favourable=is_fav,
            prior_year_actual=pcp,
            vs_pcp_abs=pcp_abs,
            vs_pcp_pct=pcp_pct,
            commentary=commentary_map.get(acct.id),
        ))

    return VarianceReportResponse(
        fy_year=fy_year,
        period_label=period_label,
        version_id=version_id,
        view_mode=view_mode,
        rows=rows,
    )


# ── PUT /reports/commentary ──────────────────────────────────────────────────


@router.put("/commentary", response_model=CommentaryRead)
async def save_commentary(
    payload: CommentaryPayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_finance),
):
    period_filter = (
        ReportCommentary.period_id == payload.period_id
        if payload.period_id is not None
        else ReportCommentary.period_id.is_(None)
    )
    result = await db.execute(
        select(ReportCommentary).where(
            ReportCommentary.version_id == payload.version_id,
            ReportCommentary.account_id == payload.account_id,
            period_filter,
        )
    )
    existing = result.scalar_one_or_none()

    now = datetime.now(timezone.utc)
    if existing:
        existing.comment = payload.comment
        existing.updated_by = user.id
        existing.updated_at = now
        await db.commit()
        await db.refresh(existing)
        return existing
    else:
        entry = ReportCommentary(
            version_id=payload.version_id,
            account_id=payload.account_id,
            period_id=payload.period_id,
            comment=payload.comment,
            updated_by=user.id,
            updated_at=now,
        )
        db.add(entry)
        await db.commit()
        await db.refresh(entry)
        return entry


# ── POST /reports/export ─────────────────────────────────────────────────────


@router.post("/export")
async def export_report(
    payload: ExportRequest,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    if payload.format != "xlsx":
        raise HTTPException(status_code=400, detail="Only xlsx format is currently supported")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    NAVY = "1F3D6E"
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    subtotal_font = Font(bold=True, size=11)
    number_fmt = "#,##0"
    pct_fmt = "0.0%"
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    # Load periods (exclude fy_month=0 opening-balance periods)
    result = await db.execute(
        select(Period)
        .where(Period.fy_year == payload.fy_year, Period.fy_month >= 1)
        .order_by(Period.fy_month)
    )
    periods = list(result.scalars().all())
    period_ids = [p.id for p in periods]
    period_labels = [_period_label(p) for p in periods]

    wb = Workbook()
    wb.remove(wb.active)

    if payload.type == "variance" and payload.version_id:
        await _export_variance_sheets(
            wb, db, payload.version_id, payload.fy_year, periods, period_ids, period_labels,
            header_font, header_fill, subtotal_font, number_fmt, pct_fmt, thin_border,
            include_aasb16=payload.include_aasb16,
        )
    elif payload.type == "budget" and payload.version_id:
        for stmt, stmt_label in [("is", "Income Statement"), ("bs", "Balance Sheet"), ("cf", "Cash Flow")]:
            await _export_budget_sheet(
                wb, db, payload.version_id, stmt, stmt_label,
                periods, period_ids, period_labels,
                header_font, header_fill, subtotal_font, number_fmt, thin_border,
            )
    elif payload.type == "actuals":
        for stmt, stmt_label in [("is", "Income Statement"), ("bs", "Balance Sheet")]:
            await _export_actuals_sheet(
                wb, db, stmt, stmt_label, period_ids, period_labels,
                header_font, header_fill, subtotal_font, number_fmt, thin_border,
                include_aasb16=payload.include_aasb16,
            )
    else:
        raise HTTPException(status_code=400, detail="Invalid export type or missing version_id")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kip_{payload.type}_FY{payload.fy_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _export_variance_sheets(wb, db, version_id, fy_year, periods, period_ids, period_labels,
                                  header_font, header_fill, subtotal_font, number_fmt, pct_fmt, thin_border,
                                  include_aasb16: bool = True):
    from openpyxl.styles import Alignment

    result = await db.execute(
        select(Account).where(Account.statement == Statement.is_).order_by(Account.sort_order)
    )
    accounts = list(result.scalars().all())
    account_ids = [a.id for a in accounts]

    actual_by_acct: dict[uuid.UUID, float] = defaultdict(float)
    result = await db.execute(
        select(ConsolidatedActual).where(
            ConsolidatedActual.period_id.in_(period_ids),
            ConsolidatedActual.is_group_total.is_(True),
            ConsolidatedActual.account_id.in_(account_ids),
        )
    )
    for act in result.scalars().all():
        actual_by_acct[act.account_id] += float(act.amount)

    if not include_aasb16 and period_ids:
        aasb16_adj = await compute_aasb16_by_account_period(db, period_ids)
        for acct_id in account_ids:
            if acct_id in aasb16_adj:
                for pid in period_ids:
                    actual_by_acct[acct_id] -= aasb16_adj[acct_id].get(pid, 0.0)

    budget_by_acct: dict[uuid.UUID, float] = defaultdict(float)
    result = await db.execute(
        select(ModelOutput).where(
            ModelOutput.version_id == version_id,
            ModelOutput.period_id.in_(period_ids),
            ModelOutput.account_id.in_(account_ids),
            ModelOutput.entity_id.is_(None),
        )
    )
    for mo in result.scalars().all():
        budget_by_acct[mo.account_id] += float(mo.amount)

    ws = wb.create_sheet(title="IS Variance")
    headers = ["Account", "Actual", "Budget", "Variance $", "Variance %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 40
    for c in ["B", "C", "D", "E"]:
        ws.column_dimensions[c].width = 14

    row_idx = 2
    for acct in accounts:
        sign = -1.0  # IS accounts stored credit-normal; flip for display
        actual = actual_by_acct.get(acct.id, 0) * sign
        budget = budget_by_acct.get(acct.id, 0) * sign
        var = actual - budget
        var_pct = var / budget if budget != 0 else None

        ws.cell(row=row_idx, column=1, value=f"{'  ' if not acct.is_subtotal else ''}{acct.name}")
        c_actual = ws.cell(row=row_idx, column=2, value=round(actual, 2))
        c_budget = ws.cell(row=row_idx, column=3, value=round(budget, 2))
        c_var = ws.cell(row=row_idx, column=4, value=round(var, 2))
        c_pct = ws.cell(row=row_idx, column=5, value=round(var_pct, 4))

        c_actual.number_format = number_fmt
        c_budget.number_format = number_fmt
        c_var.number_format = number_fmt
        c_pct.number_format = pct_fmt

        if acct.is_subtotal:
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).font = subtotal_font
                ws.cell(row=row_idx, column=col).border = thin_border

        row_idx += 1


async def _export_budget_sheet(wb, db, version_id, stmt_key, stmt_label,
                               periods, period_ids, period_labels,
                               header_font, header_fill, subtotal_font, number_fmt, thin_border):
    from openpyxl.styles import Alignment

    statement = Statement.is_ if stmt_key == "is" else (Statement.bs if stmt_key == "bs" else Statement.cf)

    if statement == Statement.cf:
        cf_codes = ["CF-OPERATING", "CF-INVESTING", "CF-FINANCING", "CF-NET"]
        result = await db.execute(select(Account).where(Account.code.in_(cf_codes)))
        accounts = list(result.scalars().all())
    else:
        result = await db.execute(
            select(Account).where(Account.statement == statement).order_by(Account.sort_order)
        )
        accounts = list(result.scalars().all())

    account_ids = [a.id for a in accounts]
    amounts: dict[tuple[uuid.UUID, uuid.UUID], float] = {}
    result = await db.execute(
        select(ModelOutput).where(
            ModelOutput.version_id == version_id,
            ModelOutput.period_id.in_(period_ids),
            ModelOutput.account_id.in_(account_ids),
            ModelOutput.entity_id.is_(None),
        )
    )
    for mo in result.scalars().all():
        amounts[(mo.account_id, mo.period_id)] = float(mo.amount)

    ws = wb.create_sheet(title=stmt_label)
    headers = ["Account"] + period_labels
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 40
    for i in range(len(period_labels)):
        ws.column_dimensions[chr(66 + i)].width = 12

    row_idx = 2
    for acct in accounts:
        ws.cell(row=row_idx, column=1, value=f"{'  ' if not acct.is_subtotal else ''}{acct.name}")
        for pi, period in enumerate(periods):
            val = amounts.get((acct.id, period.id), 0)
            c = ws.cell(row=row_idx, column=2 + pi, value=round(val, 2))
            c.number_format = number_fmt
        if acct.is_subtotal:
            for col in range(1, 2 + len(periods)):
                ws.cell(row=row_idx, column=col).font = subtotal_font
                ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1


async def _export_actuals_sheet(wb, db, stmt_key, stmt_label, period_ids, period_labels,
                                header_font, header_fill, subtotal_font, number_fmt, thin_border,
                                include_aasb16: bool = True):
    from openpyxl.styles import Alignment

    statement = Statement.is_ if stmt_key == "is" else Statement.bs
    result = await db.execute(
        select(Account).where(Account.statement == statement).order_by(Account.sort_order)
    )
    accounts = list(result.scalars().all())
    account_ids = [a.id for a in accounts]

    is_bs = statement == Statement.bs

    # For BS we need all historical periods to compute cumulative balances.
    if is_bs:
        fy_periods = await db.execute(
            select(Period).where(Period.id.in_(period_ids)).order_by(Period.fy_month)
        )
        fy_period_list = list(fy_periods.scalars().all())
        if fy_period_list:
            fy_year = fy_period_list[0].fy_year
            max_fy_month = fy_period_list[-1].fy_month
        else:
            fy_year, max_fy_month = 0, 12
        hist_result = await db.execute(
            select(Period).where(
                (Period.fy_year < fy_year)
                | ((Period.fy_year == fy_year) & (Period.fy_month <= max_fy_month))
            ).order_by(Period.fy_year, Period.fy_month)
        )
        all_hist_periods = list(hist_result.scalars().all())
        all_hist_ids = [p.id for p in all_hist_periods]
        query_period_ids = all_hist_ids
    else:
        query_period_ids = period_ids

    amounts: dict[tuple[uuid.UUID, uuid.UUID], float] = {}
    result = await db.execute(
        select(ConsolidatedActual).where(
            ConsolidatedActual.period_id.in_(query_period_ids),
            ConsolidatedActual.is_group_total.is_(True),
            ConsolidatedActual.account_id.in_(account_ids),
        )
    )
    for act in result.scalars().all():
        amounts[(act.account_id, act.period_id)] = float(act.amount)

    if not include_aasb16 and query_period_ids:
        aasb16_adj = await compute_aasb16_by_account_period(db, query_period_ids)
        for acct_id in account_ids:
            if acct_id in aasb16_adj:
                for pid in query_period_ids:
                    adj_key = (acct_id, pid)
                    adj_val = aasb16_adj[acct_id].get(pid, 0.0)
                    if adj_val != 0:
                        amounts[adj_key] = amounts.get(adj_key, 0.0) - adj_val

    # For BS, pre-compute cumulative sums at each display period.
    if is_bs:
        cum_amounts: dict[tuple[uuid.UUID, uuid.UUID], float] = {}
        for acct in accounts:
            running = 0.0
            for hp in all_hist_periods:
                running += amounts.get((acct.id, hp.id), 0.0)
                if hp.id in set(period_ids):
                    cum_amounts[(acct.id, hp.id)] = running

    ws = wb.create_sheet(title=stmt_label)
    headers = ["Account"] + period_labels
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 40
    for i in range(len(period_labels)):
        ws.column_dimensions[chr(66 + i)].width = 12

    row_idx = 2
    for acct in accounts:
        if statement == Statement.is_:
            sign = -1.0
        elif acct.normal_balance == NormalBalance.credit:
            sign = -1.0
        else:
            sign = 1.0

        ws.cell(row=row_idx, column=1, value=f"{'  ' if not acct.is_subtotal else ''}{acct.name}")
        for pi, pid in enumerate(period_ids):
            if is_bs:
                val = cum_amounts.get((acct.id, pid), 0) * sign
            else:
                val = amounts.get((acct.id, pid), 0) * sign
            c = ws.cell(row=row_idx, column=2 + pi, value=round(val, 2))
            c.number_format = number_fmt
        if acct.is_subtotal:
            for col in range(1, 2 + len(period_ids)):
                ws.cell(row=row_idx, column=col).font = subtotal_font
                ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1


# ── POST /reports/management-pack ────────────────────────────────────────

PACK_NAVY = "1F3D6E"
PACK_GREEN = "1F6E3D"
PACK_BLUE = "1F3D8E"
PACK_PURPLE = "6E1F6E"

MARGIN_CODES = {"REV-SALES", "GM", "EBITDA", "NPAT"}

BS_SECTION_HEADERS: dict[str, list[AccountType]] = {
    "Assets": [AccountType.asset],
    "Liabilities": [AccountType.liability],
    "Equity": [AccountType.equity],
}

CF_ROW_DEFS = [
    ("section", "Operating Cash Flow", None),
    ("row", "NPAT", "NPAT"),
    ("row", "Depreciation & Amortisation", "DEP-TOTAL"),
    ("row", "Working Capital Movement", None),
    ("subtotal", "Operating Cash Flow", "CF-OPERATING"),
    ("section", "Investing Cash Flow", None),
    ("row", "Capital Expenditure", "CAPEX"),
    ("subtotal", "Investing Cash Flow", "CF-INVESTING"),
    ("section", "Financing Cash Flow", None),
    ("row", "Net Debt Movement", None),
    ("subtotal", "Financing Cash Flow", "CF-FINANCING"),
    ("subtotal", "Net Cash Flow", "CF-NET"),
    ("section", "Cash Balance", None),
    ("row", "Opening Cash", None),
    ("row", "Closing Cash", "BS-CASH"),
]


def _pack_styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    return {
        "header_font": Font(color="FFFFFF", bold=True, size=10),
        "header_fill": PatternFill(start_color=PACK_NAVY, end_color=PACK_NAVY, fill_type="solid"),
        "section_font": Font(color=PACK_NAVY, bold=True, size=10),
        "subtotal_font": Font(bold=True, size=10),
        "subtotal_fill": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
        "subtotal_border": Border(top=Side(style="thin", color="999999")),
        "red_font": Font(color="CC0000", size=10),
        "red_bold_font": Font(color="CC0000", bold=True, size=10),
        "number_fmt": "#,##0",
        "pct_fmt": "0.0%",
        "align_center": Alignment(horizontal="center"),
        "align_right": Alignment(horizontal="right"),
    }


def _apply_header_row(ws, headers: list[str], styles: dict):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["align_center"]


def _set_col_widths(ws, col_count: int, margin_cols: set[int] | None = None):
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 35
    for c in range(2, col_count + 1):
        letter = get_column_letter(c)
        if margin_cols and c in margin_cols:
            ws.column_dimensions[letter].width = 8
        else:
            ws.column_dimensions[letter].width = 14


def _write_number(ws, row, col, value, styles, is_pct=False, is_subtotal=False):
    cell = ws.cell(row=row, column=col, value=round(value, 4 if is_pct else 0))
    cell.number_format = styles["pct_fmt"] if is_pct else styles["number_fmt"]
    if value < 0:
        cell.font = styles["red_bold_font"] if is_subtotal else styles["red_font"]
    elif is_subtotal:
        cell.font = styles["subtotal_font"]


async def _load_periods_for_fy(db: AsyncSession, fy_year: int) -> list[Period]:
    result = await db.execute(
        select(Period)
        .where(Period.fy_year == fy_year, Period.fy_month >= 1)
        .order_by(Period.fy_month)
    )
    return list(result.scalars().all())


async def _load_actuals_summed(
    db: AsyncSession,
    period_ids: list[uuid.UUID],
    account_ids: list[uuid.UUID],
    include_aasb16: bool,
    entity_id: uuid.UUID | None = None,
) -> dict[uuid.UUID, float]:
    """Sum ConsolidatedActual amounts across periods for given accounts."""
    totals: dict[uuid.UUID, float] = defaultdict(float)
    if not period_ids:
        return totals

    if entity_id is not None:
        result = await db.execute(
            select(ConsolidatedActual).where(
                ConsolidatedActual.period_id.in_(period_ids),
                ConsolidatedActual.entity_id == entity_id,
                ConsolidatedActual.is_group_total.is_(False),
                ConsolidatedActual.account_id.in_(account_ids),
            )
        )
    else:
        result = await db.execute(
            select(ConsolidatedActual).where(
                ConsolidatedActual.period_id.in_(period_ids),
                ConsolidatedActual.is_group_total.is_(True),
                ConsolidatedActual.account_id.in_(account_ids),
            )
        )
    for act in result.scalars().all():
        totals[act.account_id] += float(act.amount)

    if not include_aasb16:
        aasb16_adj = await compute_aasb16_by_account_period(db, period_ids)
        for acct_id in account_ids:
            if acct_id in aasb16_adj:
                for pid in period_ids:
                    totals[acct_id] -= aasb16_adj[acct_id].get(pid, 0.0)

    return totals


async def _load_budget_summed(
    db: AsyncSession,
    version_id: uuid.UUID,
    period_ids: list[uuid.UUID],
    account_ids: list[uuid.UUID],
    entity_id: uuid.UUID | None = None,
) -> dict[uuid.UUID, float]:
    totals: dict[uuid.UUID, float] = defaultdict(float)
    if not period_ids:
        return totals
    q = select(ModelOutput).where(
        ModelOutput.version_id == version_id,
        ModelOutput.period_id.in_(period_ids),
        ModelOutput.account_id.in_(account_ids),
    )
    if entity_id is not None:
        q = q.where(ModelOutput.entity_id == entity_id)
    else:
        q = q.where(ModelOutput.entity_id.is_(None))
    result = await db.execute(q)
    for mo in result.scalars().all():
        totals[mo.account_id] += float(mo.amount)
    return totals


async def _find_best_version(db: AsyncSession, fy_year: int) -> BudgetVersion | None:
    result = await db.execute(
        select(BudgetVersion).where(
            BudgetVersion.fy_year == fy_year,
            BudgetVersion.version_type == VersionType.budget,
            BudgetVersion.status == VersionStatus.approved,
        ).limit(1)
    )
    version = result.scalar_one_or_none()
    if version:
        return version
    result = await db.execute(
        select(BudgetVersion).where(
            BudgetVersion.fy_year == fy_year,
            BudgetVersion.version_type == VersionType.budget,
        ).order_by(BudgetVersion.created_at.desc()).limit(1)
    )
    return result.scalar_one_or_none()


async def _load_bs_cumulative(
    db: AsyncSession,
    fy_year: int,
    fy_month: int | None,
    account_ids: list[uuid.UUID],
    include_aasb16: bool,
    entity_id: uuid.UUID | None = None,
) -> dict[uuid.UUID, float]:
    """Load cumulative BS balances through a given point."""
    if fy_month is not None:
        q = select(Period).where(
            (Period.fy_year < fy_year)
            | ((Period.fy_year == fy_year) & (Period.fy_month <= fy_month))
        )
    else:
        q = select(Period).where(Period.fy_year <= fy_year)
    result = await db.execute(q.order_by(Period.fy_year, Period.fy_month))
    all_periods = list(result.scalars().all())
    all_period_ids = [p.id for p in all_periods]

    return await _load_actuals_summed(db, all_period_ids, account_ids, include_aasb16, entity_id)


def _display_sign_is() -> float:
    return -1.0


def _display_sign_bs(account: Account) -> float:
    if account.normal_balance == NormalBalance.credit:
        return -1.0
    return 1.0


@router.post("/management-pack")
async def generate_management_pack(
    payload: ManagementPackRequest,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    p = payload.periods
    styles = _pack_styles()

    # ── Load accounts ────────────────────────────────────────────────
    result = await db.execute(select(Account).order_by(Account.sort_order))
    all_accounts = list(result.scalars().all())
    is_accounts = [a for a in all_accounts if a.statement == Statement.is_]
    bs_accounts = [a for a in all_accounts if a.statement == Statement.bs]
    account_by_code: dict[str, Account] = {a.code: a for a in all_accounts}
    is_account_ids = [a.id for a in is_accounts]
    bs_account_ids = [a.id for a in bs_accounts]

    # ── Load entities ────────────────────────────────────────────────
    result = await db.execute(
        select(Entity).where(Entity.is_active.is_(True)).order_by(Entity.code)
    )
    entities = list(result.scalars().all())

    # ── Load periods ─────────────────────────────────────────────────
    prior2_periods = await _load_periods_for_fy(db, p.prior2_fy_year)
    prior1_periods = await _load_periods_for_fy(db, p.prior1_fy_year)
    ytd_periods = await _load_periods_for_fy(db, p.ytd_fy_year)
    ytd_period_ids = [pr.id for pr in ytd_periods if pr.fy_month <= p.ytd_to_month]
    forecast_period_ids = [pr.id for pr in ytd_periods if pr.fy_month > p.ytd_to_month]
    all_ytd_period_ids = [pr.id for pr in ytd_periods]

    # ── Resolve budget versions ──────────────────────────────────────
    budget_version: BudgetVersion | None = None
    if p.budget_version_id:
        budget_version = await db.get(BudgetVersion, p.budget_version_id)
    if budget_version is None:
        budget_version = await _find_best_version(db, p.forecast_fy_year)

    fy27_version = await _find_best_version(db, p.forecast_fy_year + 1)

    # ── Gather IS data ───────────────────────────────────────────────
    entity_filter = payload.entity_id

    prior2_actuals = await _load_actuals_summed(
        db, [pr.id for pr in prior2_periods], is_account_ids, payload.include_aasb16, entity_filter,
    )
    prior1_actuals = await _load_actuals_summed(
        db, [pr.id for pr in prior1_periods], is_account_ids, payload.include_aasb16, entity_filter,
    )
    ytd_actuals = await _load_actuals_summed(
        db, ytd_period_ids, is_account_ids, payload.include_aasb16, entity_filter,
    )

    # FYE estimate = actuals M1-ytd_to_month + forecast M(ytd_to_month+1)-12
    fye_estimate: dict[uuid.UUID, float] = defaultdict(float)
    for aid, val in ytd_actuals.items():
        fye_estimate[aid] += val

    if budget_version and forecast_period_ids:
        forecast_budget = await _load_budget_summed(
            db, budget_version.id, forecast_period_ids, is_account_ids, entity_filter,
        )
        for aid, val in forecast_budget.items():
            fye_estimate[aid] += val
    elif forecast_period_ids:
        # Trailing 3-month average as fallback
        trailing_months = min(3, p.ytd_to_month)
        trailing_pids = [
            pr.id for pr in ytd_periods
            if p.ytd_to_month - trailing_months < pr.fy_month <= p.ytd_to_month
        ]
        if trailing_pids:
            trailing_actuals = await _load_actuals_summed(
                db, trailing_pids, is_account_ids, payload.include_aasb16, entity_filter,
            )
            remaining_months = 12 - p.ytd_to_month
            for aid, val in trailing_actuals.items():
                monthly_avg = val / len(trailing_pids)
                fye_estimate[aid] += monthly_avg * remaining_months

    # Budget full year
    budget_full_year: dict[uuid.UUID, float] = {}
    if budget_version:
        budget_full_year = await _load_budget_summed(
            db, budget_version.id, all_ytd_period_ids, is_account_ids, entity_filter,
        )

    # FY2027 budget
    fy27_data: dict[uuid.UUID, float] = {}
    fy27_periods_list: list[Period] = []
    if fy27_version:
        fy27_periods_list = await _load_periods_for_fy(db, p.forecast_fy_year + 1)
        fy27_pids = [pr.id for pr in fy27_periods_list]
        if fy27_pids:
            fy27_data = await _load_budget_summed(
                db, fy27_version.id, fy27_pids, is_account_ids, entity_filter,
            )

    # ── Build workbook ───────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 1 — Income Statement
    # ═══════════════════════════════════════════════════════════════════
    ws_is = wb.create_sheet(title="Income Statement")
    ws_is.sheet_properties.tabColor = PACK_NAVY

    is_headers = [
        "Account",
        f"FY{p.prior2_fy_year}", "GM%", "EBITDA%", "NPAT%",
        f"FY{p.prior1_fy_year}", "GM%", "EBITDA%", "NPAT%",
        f"FY{p.ytd_fy_year} YTD M{p.ytd_to_month:02d}", "GM%", "EBITDA%", "NPAT%",
        f"FY{p.forecast_fy_year} FYE Est", "GM%", "EBITDA%", "NPAT%",
    ]
    if budget_version:
        is_headers += [f"FY{p.forecast_fy_year} Budget", "GM%", "EBITDA%", "NPAT%"]
    if fy27_version:
        is_headers += [f"FY{p.forecast_fy_year + 1} Budget", "GM%", "EBITDA%", "NPAT%"]
    else:
        is_headers += [f"FY{p.forecast_fy_year + 1} Budget"]

    _apply_header_row(ws_is, is_headers, styles)

    margin_cols: set[int] = set()
    col = 2
    for _ in range(4):  # prior2, prior1, ytd, fye
        col += 1; margin_cols.add(col)
        col += 1; margin_cols.add(col)
        col += 1; margin_cols.add(col)
        col += 1
    if budget_version:
        col += 1; margin_cols.add(col)
        col += 1; margin_cols.add(col)
        col += 1; margin_cols.add(col)
        col += 1
    _set_col_widths(ws_is, len(is_headers), margin_cols)

    # Pre-compute revenue for margin calculations
    rev_code_acct = account_by_code.get("REV-SALES")
    data_columns = [prior2_actuals, prior1_actuals, ytd_actuals, fye_estimate]
    if budget_version:
        data_columns.append(budget_full_year)

    rev_values = []
    for dc in data_columns:
        rev = (dc.get(rev_code_acct.id, 0.0) * _display_sign_is()) if rev_code_acct else 0.0
        rev_values.append(rev)

    # Section headers for IS
    type_to_section: dict[AccountType, str] = {}
    for sec, types in IS_SECTIONS.items():
        for t in types:
            type_to_section[t] = sec
    seen_sections: set[str] = set()

    row_idx = 2
    for acct in is_accounts:
        sec = type_to_section.get(acct.account_type)
        if sec and sec not in seen_sections:
            seen_sections.add(sec)
            ws_is.cell(row=row_idx, column=1, value=sec).font = styles["section_font"]
            row_idx += 1

        sign = _display_sign_is()
        label = acct.name
        if not acct.is_subtotal:
            label = f"  {label}"

        ws_is.cell(row=row_idx, column=1, value=label)

        col = 2
        for di, dc in enumerate(data_columns):
            val = dc.get(acct.id, 0.0) * sign
            _write_number(ws_is, row_idx, col, val, styles, is_subtotal=acct.is_subtotal)
            col += 1

            # Margin columns (GM%, EBITDA%, NPAT%)
            for margin_code in ["GM", "EBITDA", "NPAT"]:
                if acct.code in MARGIN_CODES:
                    rev = rev_values[di] if di < len(rev_values) else 0.0
                    if acct.code == margin_code and rev != 0:
                        _write_number(ws_is, row_idx, col, val / rev, styles, is_pct=True, is_subtotal=acct.is_subtotal)
                col += 1

        # Budget full year (if version exists)
        if budget_version:
            val = budget_full_year.get(acct.id, 0.0) * sign
            _write_number(ws_is, row_idx, col, val, styles, is_subtotal=acct.is_subtotal)
            col += 1
            for margin_code in ["GM", "EBITDA", "NPAT"]:
                rev = rev_values[4] if len(rev_values) > 4 else 0.0
                if acct.code == margin_code and rev != 0:
                    _write_number(ws_is, row_idx, col, val / rev, styles, is_pct=True, is_subtotal=acct.is_subtotal)
                col += 1

        # FY2027
        if fy27_version:
            val = fy27_data.get(acct.id, 0.0) * sign
            _write_number(ws_is, row_idx, col, val, styles, is_subtotal=acct.is_subtotal)
            col += 1
            # No margin % for FY27 — compute revenue for it
            fy27_rev = (fy27_data.get(rev_code_acct.id, 0.0) * sign) if rev_code_acct else 0.0
            for margin_code in ["GM", "EBITDA", "NPAT"]:
                if acct.code == margin_code and fy27_rev != 0:
                    _write_number(ws_is, row_idx, col, val / fy27_rev, styles, is_pct=True, is_subtotal=acct.is_subtotal)
                col += 1
        else:
            if row_idx == 2:
                ws_is.cell(row=row_idx, column=col, value="Budget not set")
            col += 1

        # Subtotal formatting
        if acct.is_subtotal:
            for c in range(1, len(is_headers) + 1):
                cell = ws_is.cell(row=row_idx, column=c)
                cell.font = styles["subtotal_font"] if cell.value is None or (isinstance(cell.value, (int, float)) and cell.value >= 0) else styles["red_bold_font"]
                cell.fill = styles["subtotal_fill"]
                cell.border = styles["subtotal_border"]

        row_idx += 1

    ws_is.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 2 — Balance Sheet
    # ═══════════════════════════════════════════════════════════════════
    ws_bs = wb.create_sheet(title="Balance Sheet")
    ws_bs.sheet_properties.tabColor = PACK_GREEN

    # Determine the calendar dates for column labels
    ytd_cal_month = (p.ytd_to_month + 6) if p.ytd_to_month <= 6 else (p.ytd_to_month - 6)
    ytd_cal_year = p.ytd_fy_year if p.ytd_to_month > 6 else p.ytd_fy_year - 1

    bs_headers = [
        "Account",
        f"30 Jun {p.prior2_fy_year}",
        f"30 Jun {p.prior1_fy_year}",
        f"{ytd_cal_month:02d}/{ytd_cal_year} (YTD)",
    ]
    has_bs_forecast = budget_version is not None
    if has_bs_forecast:
        bs_headers.append(f"FY{p.forecast_fy_year} FYE Forecast")

    _apply_header_row(ws_bs, bs_headers, styles)
    _set_col_widths(ws_bs, len(bs_headers))

    # Load cumulative BS at each point
    prior2_bs = await _load_bs_cumulative(
        db, p.prior2_fy_year, 12, bs_account_ids, payload.include_aasb16, entity_filter,
    )
    prior1_bs = await _load_bs_cumulative(
        db, p.prior1_fy_year, 12, bs_account_ids, payload.include_aasb16, entity_filter,
    )
    ytd_bs = await _load_bs_cumulative(
        db, p.ytd_fy_year, p.ytd_to_month, bs_account_ids, payload.include_aasb16, entity_filter,
    )

    forecast_bs: dict[uuid.UUID, float] = {}
    if has_bs_forecast:
        # FYE forecast: cumulative actuals through ytd + budget movements for remaining months
        fye_bs_periods = [pr.id for pr in ytd_periods if pr.fy_month == 12]
        if fye_bs_periods and budget_version:
            forecast_bs = await _load_budget_summed(
                db, budget_version.id, fye_bs_periods, bs_account_ids, entity_filter,
            )
            # Add actuals through ytd
            for aid in bs_account_ids:
                forecast_bs[aid] = ytd_bs.get(aid, 0.0) + forecast_bs.get(aid, 0.0)

    bs_section_map: dict[AccountType, str] = {}
    for sec, types in BS_SECTION_HEADERS.items():
        for t in types:
            bs_section_map[t] = sec
    seen_bs_sections: set[str] = set()

    row_idx = 2
    for acct in bs_accounts:
        sec = bs_section_map.get(acct.account_type)
        if sec and sec not in seen_bs_sections:
            seen_bs_sections.add(sec)
            ws_bs.cell(row=row_idx, column=1, value=sec).font = styles["section_font"]
            row_idx += 1

        sign = _display_sign_bs(acct)
        label = acct.name if acct.is_subtotal else f"  {acct.name}"
        ws_bs.cell(row=row_idx, column=1, value=label)

        bs_data_cols = [prior2_bs, prior1_bs, ytd_bs]
        if has_bs_forecast:
            bs_data_cols.append(forecast_bs)

        for ci, dc in enumerate(bs_data_cols):
            val = dc.get(acct.id, 0.0) * sign
            _write_number(ws_bs, row_idx, ci + 2, val, styles, is_subtotal=acct.is_subtotal)

        if acct.is_subtotal:
            for c in range(1, len(bs_headers) + 1):
                cell = ws_bs.cell(row=row_idx, column=c)
                cell.fill = styles["subtotal_fill"]
                cell.border = styles["subtotal_border"]
                if not (isinstance(cell.value, (int, float)) and cell.value < 0):
                    cell.font = styles["subtotal_font"]

        row_idx += 1

    ws_bs.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 3 — Cash Flow
    # ═══════════════════════════════════════════════════════════════════
    ws_cf = wb.create_sheet(title="Cash Flow")
    ws_cf.sheet_properties.tabColor = PACK_BLUE

    cf_headers = [
        "Account",
        f"FY{p.prior2_fy_year}",
        f"FY{p.prior1_fy_year}",
        f"FY{p.ytd_fy_year} YTD M{p.ytd_to_month:02d}",
        f"FY{p.forecast_fy_year} FYE Est",
    ]
    _apply_header_row(ws_cf, cf_headers, styles)
    _set_col_widths(ws_cf, len(cf_headers))

    # Load CF account codes that exist
    cf_relevant_codes = [
        "NPAT", "DEP-TOTAL", "CF-OPERATING", "CF-INVESTING", "CF-FINANCING",
        "CF-NET", "BS-CASH", "CAPEX",
    ]
    cf_accounts_map: dict[str, Account] = {}
    for code in cf_relevant_codes:
        if code in account_by_code:
            cf_accounts_map[code] = account_by_code[code]

    # For IS-based CF items (NPAT, DEP), use actuals sign-flipped.
    # For BS items (BS-CASH), use cumulative BS logic.
    # Collect all needed account IDs.
    cf_all_ids = [a.id for a in cf_accounts_map.values()]

    # Actuals for each year
    cf_prior2 = await _load_actuals_summed(
        db, [pr.id for pr in prior2_periods], cf_all_ids, payload.include_aasb16, entity_filter,
    )
    cf_prior1 = await _load_actuals_summed(
        db, [pr.id for pr in prior1_periods], cf_all_ids, payload.include_aasb16, entity_filter,
    )
    cf_ytd = await _load_actuals_summed(
        db, ytd_period_ids, cf_all_ids, payload.include_aasb16, entity_filter,
    )
    cf_fye = dict(fye_estimate)
    # Add any CF-specific accounts from budget that aren't IS accounts
    if budget_version and forecast_period_ids:
        cf_forecast = await _load_budget_summed(
            db, budget_version.id, forecast_period_ids, cf_all_ids, entity_filter,
        )
        for aid, val in cf_forecast.items():
            if aid not in [a.id for a in is_accounts]:
                cf_fye[aid] = cf_fye.get(aid, 0.0) + val

    cf_data_cols = [cf_prior2, cf_prior1, cf_ytd, cf_fye]

    # For BS-CASH, we need cumulative balances
    cash_acct = cf_accounts_map.get("BS-CASH")
    cash_cumulative = []
    if cash_acct:
        for fy, fm in [
            (p.prior2_fy_year, 12),
            (p.prior1_fy_year, 12),
            (p.ytd_fy_year, p.ytd_to_month),
            (p.forecast_fy_year, 12),
        ]:
            cum = await _load_bs_cumulative(db, fy, fm, [cash_acct.id], payload.include_aasb16, entity_filter)
            cash_cumulative.append(cum.get(cash_acct.id, 0.0))

    row_idx = 2
    for row_type, label, code in CF_ROW_DEFS:
        if row_type == "section":
            ws_cf.cell(row=row_idx, column=1, value=label).font = styles["section_font"]
            row_idx += 1
            continue

        is_sub = row_type == "subtotal"
        display_label = label if is_sub else f"  {label}"
        ws_cf.cell(row=row_idx, column=1, value=display_label)

        acct = cf_accounts_map.get(code) if code else None

        for ci, dc in enumerate(cf_data_cols):
            if code == "BS-CASH" and cash_cumulative:
                val = cash_cumulative[ci] if ci < len(cash_cumulative) else 0.0
            elif acct:
                val = dc.get(acct.id, 0.0)
                # IS items need sign flip; BS/CF items may not
                if acct.statement == Statement.is_:
                    val *= _display_sign_is()
            else:
                val = 0.0

            _write_number(ws_cf, row_idx, ci + 2, val, styles, is_subtotal=is_sub)

        if is_sub:
            for c in range(1, len(cf_headers) + 1):
                cell = ws_cf.cell(row=row_idx, column=c)
                cell.fill = styles["subtotal_fill"]
                cell.border = styles["subtotal_border"]
                if not (isinstance(cell.value, (int, float)) and cell.value < 0):
                    cell.font = styles["subtotal_font"]

        row_idx += 1

    ws_cf.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 4 — Entity Summary (consolidated only)
    # ═══════════════════════════════════════════════════════════════════
    if entity_filter is None:
        ws_ent = wb.create_sheet(title="Entity Summary")
        ws_ent.sheet_properties.tabColor = PACK_PURPLE

        key_codes = ["REV-SALES", "GM", "EBITDA", "NPAT"]
        key_accts = [account_by_code[c] for c in key_codes if c in account_by_code]
        key_ids = [a.id for a in key_accts]

        ent_headers = ["Entity"]
        for yr_label in [f"FY{p.prior1_fy_year}", f"FY{p.ytd_fy_year} YTD"]:
            for metric in ["Revenue", "GM", "EBITDA", "NPAT"]:
                ent_headers.append(f"{yr_label} {metric}")
        _apply_header_row(ws_ent, ent_headers, styles)
        _set_col_widths(ws_ent, len(ent_headers))

        entity_rows = []
        for ent in entities:
            ent_prior1 = await _load_actuals_summed(
                db, [pr.id for pr in prior1_periods], key_ids, payload.include_aasb16, ent.id,
            )
            ent_ytd = await _load_actuals_summed(
                db, ytd_period_ids, key_ids, payload.include_aasb16, ent.id,
            )
            rev_val = ent_prior1.get(key_accts[0].id, 0.0) * _display_sign_is() if key_accts else 0.0
            entity_rows.append((ent, ent_prior1, ent_ytd, rev_val))

        entity_rows.sort(key=lambda x: x[3], reverse=True)

        row_idx = 2
        for ent, ent_prior1, ent_ytd, _ in entity_rows:
            ws_ent.cell(row=row_idx, column=1, value=f"{ent.code} — {ent.name or ent.code}")
            col = 2
            for dc in [ent_prior1, ent_ytd]:
                for ka in key_accts:
                    val = dc.get(ka.id, 0.0) * _display_sign_is()
                    _write_number(ws_ent, row_idx, col, val, styles)
                    col += 1
            row_idx += 1

        ws_ent.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 5 — Assumptions (when budget version provided)
    # ═══════════════════════════════════════════════════════════════════
    if budget_version:
        ws_assum = wb.create_sheet(title="Assumptions")
        ws_assum.sheet_properties.tabColor = PACK_NAVY

        assum_headers = ["Assumption", "Entity", "Value"]
        _apply_header_row(ws_assum, assum_headers, styles)
        _set_col_widths(ws_assum, len(assum_headers))

        result = await db.execute(
            select(ModelAssumption).where(
                ModelAssumption.budget_version_id == budget_version.id,
            ).order_by(ModelAssumption.assumption_key)
        )
        assumptions = list(result.scalars().all())

        entity_map = {e.id: e for e in entities}
        row_idx = 2
        for assum in assumptions:
            ws_assum.cell(row=row_idx, column=1, value=assum.assumption_key)
            ent = entity_map.get(assum.entity_id) if assum.entity_id else None
            ws_assum.cell(row=row_idx, column=2, value=ent.code if ent else "Group")

            val = assum.assumption_value
            if isinstance(val, dict):
                display_val = ", ".join(f"{k}: {v}" for k, v in val.items())
            else:
                display_val = str(val)
            ws_assum.cell(row=row_idx, column=3, value=display_val)
            row_idx += 1

        ws_assum.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════════════
    # PER-ENTITY SHEETS — Standalone IS + BS for each entity
    # ═══════════════════════════════════════════════════════════════════
    if entity_filter is None:
        for ent in entities:
            # ── Entity IS ──
            ent_is_name = f"{ent.code} IS"
            ws_ent_is = wb.create_sheet(title=ent_is_name[:31])

            ent_is_headers = [
                "Account",
                f"FY{p.prior2_fy_year}", f"FY{p.prior1_fy_year}",
                f"FY{p.ytd_fy_year} YTD",
            ]
            _apply_header_row(ws_ent_is, ent_is_headers, styles)
            _set_col_widths(ws_ent_is, len(ent_is_headers))

            ent_p2 = await _load_actuals_summed(
                db, [pr.id for pr in prior2_periods], is_account_ids, payload.include_aasb16, ent.id,
            )
            ent_p1 = await _load_actuals_summed(
                db, [pr.id for pr in prior1_periods], is_account_ids, payload.include_aasb16, ent.id,
            )
            ent_ytd = await _load_actuals_summed(
                db, ytd_period_ids, is_account_ids, payload.include_aasb16, ent.id,
            )

            seen = set()
            r = 2
            for acct in is_accounts:
                sec = type_to_section.get(acct.account_type)
                if sec and sec not in seen:
                    seen.add(sec)
                    ws_ent_is.cell(row=r, column=1, value=sec).font = styles["section_font"]
                    r += 1

                lbl = acct.name if acct.is_subtotal else f"  {acct.name}"
                ws_ent_is.cell(row=r, column=1, value=lbl)
                for ci, dc in enumerate([ent_p2, ent_p1, ent_ytd]):
                    val = dc.get(acct.id, 0.0) * _display_sign_is()
                    _write_number(ws_ent_is, r, ci + 2, val, styles, is_subtotal=acct.is_subtotal)

                if acct.is_subtotal:
                    for c in range(1, len(ent_is_headers) + 1):
                        cell = ws_ent_is.cell(row=r, column=c)
                        cell.fill = styles["subtotal_fill"]
                        cell.border = styles["subtotal_border"]
                        if not (isinstance(cell.value, (int, float)) and cell.value < 0):
                            cell.font = styles["subtotal_font"]
                r += 1

            ws_ent_is.freeze_panes = "B2"

            # ── Entity BS ──
            ent_bs_name = f"{ent.code} BS"
            ws_ent_bs = wb.create_sheet(title=ent_bs_name[:31])

            ent_bs_headers = [
                "Account",
                f"30 Jun {p.prior2_fy_year}", f"30 Jun {p.prior1_fy_year}",
                f"YTD M{p.ytd_to_month:02d}",
            ]
            _apply_header_row(ws_ent_bs, ent_bs_headers, styles)
            _set_col_widths(ws_ent_bs, len(ent_bs_headers))

            ent_bs_p2 = await _load_bs_cumulative(
                db, p.prior2_fy_year, 12, bs_account_ids, payload.include_aasb16, ent.id,
            )
            ent_bs_p1 = await _load_bs_cumulative(
                db, p.prior1_fy_year, 12, bs_account_ids, payload.include_aasb16, ent.id,
            )
            ent_bs_ytd = await _load_bs_cumulative(
                db, p.ytd_fy_year, p.ytd_to_month, bs_account_ids, payload.include_aasb16, ent.id,
            )

            seen_bs = set()
            r = 2
            for acct in bs_accounts:
                sec = bs_section_map.get(acct.account_type)
                if sec and sec not in seen_bs:
                    seen_bs.add(sec)
                    ws_ent_bs.cell(row=r, column=1, value=sec).font = styles["section_font"]
                    r += 1

                sign = _display_sign_bs(acct)
                lbl = acct.name if acct.is_subtotal else f"  {acct.name}"
                ws_ent_bs.cell(row=r, column=1, value=lbl)
                for ci, dc in enumerate([ent_bs_p2, ent_bs_p1, ent_bs_ytd]):
                    val = dc.get(acct.id, 0.0) * sign
                    _write_number(ws_ent_bs, r, ci + 2, val, styles, is_subtotal=acct.is_subtotal)

                if acct.is_subtotal:
                    for c in range(1, len(ent_bs_headers) + 1):
                        cell = ws_ent_bs.cell(row=r, column=c)
                        cell.fill = styles["subtotal_fill"]
                        cell.border = styles["subtotal_border"]
                        if not (isinstance(cell.value, (int, float)) and cell.value < 0):
                            cell.font = styles["subtotal_font"]
                r += 1

            ws_ent_bs.freeze_panes = "B2"

    # ── Write output ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    month_names = [
        "", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    ]
    m_name = month_names[p.ytd_to_month] if 1 <= p.ytd_to_month <= 12 else "Unknown"
    filename = f"KipGroup_ManagementPack_{m_name}{p.ytd_fy_year}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/management-pack/versions")
async def list_budget_versions_for_pack(
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_finance),
):
    """Return available budget versions for the management pack modal."""
    result = await db.execute(
        select(BudgetVersion)
        .where(BudgetVersion.version_type == VersionType.budget)
        .order_by(BudgetVersion.fy_year.desc(), BudgetVersion.created_at.desc())
    )
    versions = result.scalars().all()
    return [
        {
            "id": str(v.id),
            "name": v.name,
            "fy_year": v.fy_year,
            "status": v.status.value if v.status else "draft",
        }
        for v in versions
    ]
