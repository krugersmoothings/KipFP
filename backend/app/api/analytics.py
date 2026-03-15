"""Analytics endpoints — time series and location performance."""

from __future__ import annotations

import io
import uuid
from collections import defaultdict

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_db, require_viewer
from app.db.models.account import Account, AccountMapping, AccountType
from app.db.models.budget import BudgetVersion, ModelOutput
from app.db.models.consolidation import ConsolidatedActual
from app.db.models.entity import Entity
from app.db.models.location import Location
from app.db.models.period import Period
from app.db.models.sync import JeLine
from app.db.models.user import User
from app.services.aasb16_helpers import compute_aasb16_by_account_period
from app.schemas.analytics import (
    AnalyticsExportRequest,
    LocationPerformanceRow,
    LocationTimeSeriesPoint,
    MultiTimeSeriesResponse,
    MultiTimeSeriesSeries,
    TimeSeriesPoint,
)

router = APIRouter(prefix="/analytics", tags=["analytics"])

METRIC_CODES = {
    "revenue": "REV-SALES",
    "gm": "GM",
    "ebitda": "EBITDA",
    "npat": "NPAT",
}

DIRECT_COST_CODES = {"OPEX-WAGES", "OPEX-SUPER", "OPEX-PAYROLLTAX", "COGS"}

# All four KPI metrics are credit-normal in the COA, stored as negative in the DB.
# Negate them so charts/tables show positive values for favourable results.
_NEGATE_METRICS = True

MONTH_ABBR = [
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
]


async def _resolve_aasb16_for_account(
    db: AsyncSession,
    account: Account,
    aasb16_adj: dict[uuid.UUID, dict[uuid.UUID, float]],
) -> dict[uuid.UUID, float]:
    """Resolve the total AASB16 adjustment for a (possibly subtotal) account.

    For leaf accounts, returns the adjustment directly from aasb16_adj.
    For subtotals, recursively expands the formula tree and sums leaf
    adjustments.  IS accounts ADD all formula items (add + subtract)
    because debit-positive convention already encodes the sign — this
    matches the consolidation engine behaviour.
    """
    if not account.is_subtotal or not account.subtotal_formula:
        return aasb16_adj.get(account.id, {})

    all_accounts_result = await db.execute(select(Account))
    acct_by_code: dict[str, Account] = {a.code: a for a in all_accounts_result.scalars().all()}

    is_pl = account.statement and account.statement.value == "is"
    result: dict[uuid.UUID, float] = defaultdict(float)

    def _collect(formula: dict, sign: float) -> None:
        for code in formula.get("add", []):
            a = acct_by_code.get(code)
            if not a:
                continue
            if a.is_subtotal and a.subtotal_formula:
                _collect(a.subtotal_formula, sign)
            else:
                for pid, amt in aasb16_adj.get(a.id, {}).items():
                    result[pid] += amt * sign

        for code in formula.get("subtract", []):
            a = acct_by_code.get(code)
            if not a:
                continue
            sub_sign = sign if is_pl else -sign
            if a.is_subtotal and a.subtotal_formula:
                _collect(a.subtotal_formula, sub_sign)
            else:
                for pid, amt in aasb16_adj.get(a.id, {}).items():
                    result[pid] += amt * sub_sign

    _collect(account.subtotal_formula, 1.0)
    return dict(result)


def _period_label(p: Period) -> str:
    if p.period_start:
        return p.period_start.strftime("%b-%y")
    cal_year = p.fy_year - 1 if p.fy_month <= 6 else p.fy_year
    return f"{MONTH_ABBR[p.fy_month - 1]}-{cal_year % 100:02d}"


def _period_sort_key(p: Period) -> tuple[int, int]:
    return (p.fy_year, p.fy_month)


def _add_rolling_averages(points: list[dict], value_key: str = "value") -> None:
    """Mutate a list of point dicts to add rolling 3M and 12M averages."""
    for i, pt in enumerate(points):
        vals_3 = [points[j][value_key] for j in range(max(0, i - 2), i + 1)]
        pt["rolling_3m_avg"] = round(sum(vals_3) / len(vals_3), 2) if vals_3 else None

        vals_12 = [points[j][value_key] for j in range(max(0, i - 11), i + 1)]
        pt["rolling_12m_avg"] = round(sum(vals_12) / len(vals_12), 2) if vals_12 else None


def _add_mom_change(points: list[dict], value_key: str = "value") -> None:
    """Mutate a list of point dicts to add month-on-month change %."""
    for i, pt in enumerate(points):
        if i == 0:
            pt["mom_change_pct"] = None
        else:
            prev = points[i - 1][value_key]
            pt["mom_change_pct"] = round((pt[value_key] - prev) / prev * 100, 2) if prev != 0 else None


async def _get_period_range(
    db: AsyncSession,
    from_fy_year: int,
    from_fy_month: int,
    to_fy_year: int,
    to_fy_month: int,
) -> list[Period]:
    """Return periods in the given range, sorted chronologically."""
    result = await db.execute(
        select(Period).where(
            # FIX(M9): exclude fy_month=0 opening balance periods
            Period.fy_month >= 1,
            ((Period.fy_year > from_fy_year) |
             ((Period.fy_year == from_fy_year) & (Period.fy_month >= from_fy_month))),
            ((Period.fy_year < to_fy_year) |
             ((Period.fy_year == to_fy_year) & (Period.fy_month <= to_fy_month))),
        ).order_by(Period.fy_year, Period.fy_month)
    )
    return list(result.scalars().all())


async def _get_prior_year_periods(
    db: AsyncSession, periods: list[Period]
) -> dict[tuple[int, int], Period]:
    """For each period, find the prior-year equivalent. Returns {(fy_year, fy_month): period}."""
    if not periods:
        return {}
    prior_fy_years = {p.fy_year - 1 for p in periods}
    fy_months = {p.fy_month for p in periods}
    result = await db.execute(
        select(Period).where(
            Period.fy_year.in_(prior_fy_years),
            Period.fy_month.in_(fy_months),
        )
    )
    return {(p.fy_year, p.fy_month): p for p in result.scalars().all()}


# ── GET /analytics/timeseries ────────────────────────────────────────────────


@router.get("/timeseries", response_model=list[TimeSeriesPoint])
async def get_timeseries(
    metric: str = Query(..., description="revenue, gm, ebitda, or npat"),
    from_fy_year: int = Query(...),
    from_fy_month: int = Query(...),
    to_fy_year: int = Query(...),
    to_fy_month: int = Query(...),
    entity_ids: str | None = Query(None, description="Comma-separated entity UUIDs"),
    include_aasb16: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_viewer),
):
    account_code = METRIC_CODES.get(metric)
    if not account_code:
        raise HTTPException(status_code=400, detail=f"Unknown metric: {metric}. Use: {list(METRIC_CODES.keys())}")

    acct_result = await db.execute(select(Account).where(Account.code == account_code))
    account = acct_result.scalar_one_or_none()
    if account is None:
        raise HTTPException(status_code=404, detail=f"Account {account_code} not found")

    periods = await _get_period_range(db, from_fy_year, from_fy_month, to_fy_year, to_fy_month)
    if not periods:
        return []

    period_ids = [p.id for p in periods]

    # Parse entity filter
    entity_filter_ids: list[uuid.UUID] | None = None
    if entity_ids:
        entity_filter_ids = [uuid.UUID(eid.strip()) for eid in entity_ids.split(",") if eid.strip()]

    # Fetch actuals
    q = select(ConsolidatedActual).where(
        ConsolidatedActual.period_id.in_(period_ids),
        ConsolidatedActual.account_id == account.id,
    )
    if entity_filter_ids:
        q = q.where(ConsolidatedActual.entity_id.in_(entity_filter_ids))
    else:
        q = q.where(ConsolidatedActual.is_group_total.is_(True))

    result = await db.execute(q)
    sign = -1.0 if _NEGATE_METRICS else 1.0
    amount_by_period: dict[uuid.UUID, float] = defaultdict(float)
    for row in result.scalars().all():
        amount_by_period[row.period_id] += float(row.amount) * sign

    # Prior year data
    prior_map = await _get_prior_year_periods(db, periods)
    prior_period_ids = [p.id for p in prior_map.values()]

    prior_amounts: dict[uuid.UUID, float] = defaultdict(float)
    if prior_period_ids:
        q_prior = select(ConsolidatedActual).where(
            ConsolidatedActual.period_id.in_(prior_period_ids),
            ConsolidatedActual.account_id == account.id,
        )
        if entity_filter_ids:
            q_prior = q_prior.where(ConsolidatedActual.entity_id.in_(entity_filter_ids))
        else:
            q_prior = q_prior.where(ConsolidatedActual.is_group_total.is_(True))
        result = await db.execute(q_prior)
        for row in result.scalars().all():
            prior_amounts[row.period_id] += float(row.amount) * sign

    # AASB16 adjustment
    aasb16_adj: dict[uuid.UUID, dict[uuid.UUID, float]] = {}
    resolved_adj: dict[uuid.UUID, float] = {}
    if not include_aasb16:
        aasb16_adj = await compute_aasb16_by_account_period(db, period_ids)
        resolved_adj = await _resolve_aasb16_for_account(db, account, aasb16_adj)

    # region agent log
    import json as _json, pathlib as _pathlib
    _log_path = _pathlib.Path("debug-70c8ca.log")
    _resolved_sample = {str(k): v for k, v in list(resolved_adj.items())[:3]} if resolved_adj else {}
    _log_entry = _json.dumps({"sessionId":"70c8ca","location":"analytics.py:timeseries","message":"aasb16 post-fix","data":{"include_aasb16":include_aasb16,"metric":metric,"account_code":account_code,"is_subtotal":account.is_subtotal,"resolved_adj_total":sum(resolved_adj.values()),"resolved_adj_count":len(resolved_adj),"resolved_sample":_resolved_sample},"timestamp":__import__("time").time(),"hypothesisId":"A-fix"})
    with open(_log_path, "a") as _f: _f.write(_log_entry + "\n")
    # endregion

    # Build points
    points: list[dict] = []
    for p in periods:
        val = amount_by_period.get(p.id, 0.0)
        if not include_aasb16 and resolved_adj:
            val -= resolved_adj.get(p.id, 0.0) * sign
        prior_key = (p.fy_year - 1, p.fy_month)
        prior_period = prior_map.get(prior_key)
        prior_val = prior_amounts.get(prior_period.id, 0.0) if prior_period else None
        points.append({
            "period_label": _period_label(p),
            "fy_year": p.fy_year,
            "fy_month": p.fy_month,
            "value": round(val, 2),
            "prior_year_value": round(prior_val, 2) if prior_val is not None else None,
        })

    _add_mom_change(points)
    _add_rolling_averages(points)

    return [TimeSeriesPoint(**pt) for pt in points]


# ── GET /analytics/timeseries/multi ──────────────────────────────────────────


@router.get("/timeseries/multi", response_model=MultiTimeSeriesResponse)
async def get_timeseries_multi(
    metrics: str = Query(..., description="Comma-separated: revenue,gm,ebitda,npat"),
    from_fy_year: int = Query(...),
    from_fy_month: int = Query(...),
    to_fy_year: int = Query(...),
    to_fy_month: int = Query(...),
    entity_ids: str | None = Query(None),
    include_aasb16: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_viewer),
):
    metric_list = [m.strip() for m in metrics.split(",") if m.strip()]
    invalid = [m for m in metric_list if m not in METRIC_CODES]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Unknown metrics: {invalid}")

    periods = await _get_period_range(db, from_fy_year, from_fy_month, to_fy_year, to_fy_month)
    if not periods:
        return MultiTimeSeriesResponse(periods=[], series=[])

    period_ids = [p.id for p in periods]
    period_labels = [_period_label(p) for p in periods]

    entity_filter_ids: list[uuid.UUID] | None = None
    if entity_ids:
        entity_filter_ids = [uuid.UUID(eid.strip()) for eid in entity_ids.split(",") if eid.strip()]

    # Load all requested accounts
    codes = [METRIC_CODES[m] for m in metric_list]
    acct_result = await db.execute(select(Account).where(Account.code.in_(codes)))
    accounts = {a.code: a for a in acct_result.scalars().all()}

    # Fetch all actuals in one query
    account_ids = [a.id for a in accounts.values()]
    q = select(ConsolidatedActual).where(
        ConsolidatedActual.period_id.in_(period_ids),
        ConsolidatedActual.account_id.in_(account_ids),
    )
    if entity_filter_ids:
        q = q.where(ConsolidatedActual.entity_id.in_(entity_filter_ids))
    else:
        q = q.where(ConsolidatedActual.is_group_total.is_(True))

    result = await db.execute(q)
    sign = -1.0 if _NEGATE_METRICS else 1.0
    # {account_id: {period_id: amount}}
    amounts: dict[uuid.UUID, dict[uuid.UUID, float]] = defaultdict(lambda: defaultdict(float))
    for row in result.scalars().all():
        amounts[row.account_id][row.period_id] += float(row.amount) * sign

    aasb16_adj: dict[uuid.UUID, dict[uuid.UUID, float]] = {}
    resolved_adjs: dict[str, dict[uuid.UUID, float]] = {}
    if not include_aasb16:
        aasb16_adj = await compute_aasb16_by_account_period(db, period_ids)
        for code, acct in accounts.items():
            resolved_adjs[code] = await _resolve_aasb16_for_account(db, acct, aasb16_adj)

    series: list[MultiTimeSeriesSeries] = []
    for m in metric_list:
        code = METRIC_CODES[m]
        acct = accounts.get(code)
        if not acct:
            series.append(MultiTimeSeriesSeries(metric=m, values=[0.0] * len(periods)))
            continue
        r_adj = resolved_adjs.get(code, {})
        vals: list[float] = []
        for p in periods:
            v = amounts[acct.id].get(p.id, 0.0)
            if not include_aasb16 and r_adj:
                v -= r_adj.get(p.id, 0.0) * sign
            vals.append(round(v, 2))
        series.append(MultiTimeSeriesSeries(metric=m, values=vals))

    return MultiTimeSeriesResponse(periods=period_labels, series=series)


# ── GET /analytics/locations ─────────────────────────────────────────────────


@router.get("/locations", response_model=list[LocationPerformanceRow])
async def get_location_performance(
    fy_year: int = Query(...),
    fy_month: int | None = Query(None, description="Omit for full year"),
    version_id: uuid.UUID | None = Query(None, description="Budget version for comparison"),
    include_aasb16: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_viewer),
):
    # Determine periods (exclude fy_month=0 opening balance periods)
    if fy_month is not None:
        result = await db.execute(
            select(Period).where(
                Period.fy_year == fy_year,
                Period.fy_month == fy_month,
                Period.fy_month >= 1,
            )
        )
        target_periods = list(result.scalars().all())
    else:
        result = await db.execute(
            select(Period).where(
                Period.fy_year == fy_year,
                Period.fy_month >= 1,
            ).order_by(Period.fy_month)
        )
        target_periods = list(result.scalars().all())

    if not target_periods:
        return []

    period_ids = [p.id for p in target_periods]

    # Load locations and entities
    loc_result = await db.execute(select(Location).where(Location.is_active.is_(True)))
    locations = {loc.id: loc for loc in loc_result.scalars().all()}

    ent_result = await db.execute(select(Entity))
    entities = {e.id: e for e in ent_result.scalars().all()}

    # Load target accounts for revenue and direct costs
    rev_result = await db.execute(select(Account).where(Account.code == "REV-SALES"))
    rev_account = rev_result.scalar_one_or_none()

    dc_result = await db.execute(select(Account).where(Account.code.in_(DIRECT_COST_CODES)))
    dc_accounts = {a.code: a for a in dc_result.scalars().all()}

    revenue_account_ids = {rev_account.id} if rev_account else set()
    direct_cost_account_ids = {a.id for a in dc_accounts.values()}
    all_target_ids = revenue_account_ids | direct_cost_account_ids

    # Load account mappings to identify source codes → target accounts
    mapping_result = await db.execute(
        select(AccountMapping).where(AccountMapping.target_account_id.in_(all_target_ids))
    )
    mappings = list(mapping_result.scalars().all())

    # Build lookup: (entity_id, source_account_code) → (target_account_id, multiplier)
    source_to_target: dict[tuple[uuid.UUID, str], tuple[uuid.UUID, float]] = {}
    for m in mappings:
        source_to_target[(m.entity_id, m.source_account_code)] = (m.target_account_id, float(m.multiplier))

    # Fetch je_lines with location_id set
    je_q = select(JeLine).where(
        JeLine.period_id.in_(period_ids),
        JeLine.location_id.isnot(None),
    )
    if not include_aasb16:
        je_q = je_q.where(JeLine.is_aasb16.is_(False))
    je_result = await db.execute(je_q)
    je_lines = list(je_result.scalars().all())

    # Aggregate by location
    # {location_id: {"revenue": float, "direct_costs": float}}
    loc_actuals: dict[uuid.UUID, dict[str, float]] = defaultdict(lambda: {"revenue": 0.0, "direct_costs": 0.0})

    for jl in je_lines:
        key = (jl.entity_id, jl.source_account_code)
        mapping = source_to_target.get(key)
        if not mapping:
            continue
        target_id, multiplier = mapping
        amount = float(jl.amount) * multiplier

        if target_id in revenue_account_ids:
            # Revenue is credit-normal (negative in DB) — negate to positive
            loc_actuals[jl.location_id]["revenue"] += -amount
        elif target_id in direct_cost_account_ids:
            loc_actuals[jl.location_id]["direct_costs"] += amount

    # Budget comparison (if version selected)
    budget_by_loc: dict[uuid.UUID, dict[str, float]] = {}
    if version_id:
        version = await db.get(BudgetVersion, version_id)
        if version:
            from app.db.models.location import SiteBudgetEntry
            budget_result = await db.execute(
                select(SiteBudgetEntry).where(
                    SiteBudgetEntry.version_id == version_id,
                    SiteBudgetEntry.location_id.isnot(None),
                )
            )
            for entry in budget_result.scalars().all():
                if entry.location_id not in budget_by_loc:
                    budget_by_loc[entry.location_id] = {"revenue": 0.0, "direct_costs": 0.0}
                item = (entry.model_line_item or "").lower()
                amt = float(entry.amount or 0)
                if "revenue" in item or "sales" in item:
                    budget_by_loc[entry.location_id]["revenue"] += amt
                elif any(k in item for k in ["wages", "super", "payroll", "cogs", "cost"]):
                    budget_by_loc[entry.location_id]["direct_costs"] += amt

    # Build response rows
    rows: list[LocationPerformanceRow] = []
    all_location_ids = set(loc_actuals.keys())
    for loc_id in all_location_ids:
        loc = locations.get(loc_id)
        if not loc:
            continue

        entity = entities.get(loc.entity_id) if loc.entity_id else None
        actuals = loc_actuals[loc_id]
        revenue = round(actuals["revenue"], 2)
        direct_costs = round(actuals["direct_costs"], 2)
        site_pl = round(revenue - direct_costs, 2)

        budget_data = budget_by_loc.get(loc_id)
        budget_rev = round(budget_data["revenue"], 2) if budget_data else None
        budget_dc = round(budget_data["direct_costs"], 2) if budget_data else None
        budget_pl = round(budget_rev - budget_dc, 2) if budget_data and budget_rev is not None and budget_dc is not None else None

        var_abs = round(site_pl - budget_pl, 2) if budget_pl is not None else None
        var_pct = round(var_abs / abs(budget_pl) * 100, 2) if budget_pl and var_abs is not None else None
        is_fav = var_abs > 0 if var_abs is not None else None

        rows.append(LocationPerformanceRow(
            location_id=loc_id,
            location_code=loc.code,
            location_name=loc.name,
            state=loc.state,
            entity_code=entity.code if entity else None,
            revenue=revenue,
            direct_costs=direct_costs,
            site_pl=site_pl,
            budget_revenue=budget_rev,
            budget_direct_costs=budget_dc,
            budget_site_pl=budget_pl,
            variance_abs=var_abs,
            variance_pct=var_pct,
            is_favourable=is_fav,
        ))

    rows.sort(key=lambda r: r.site_pl, reverse=True)
    return rows


# ── GET /analytics/locations/{location_id}/timeseries ────────────────────────


@router.get("/locations/{location_id}/timeseries", response_model=list[LocationTimeSeriesPoint])
async def get_location_timeseries(
    location_id: uuid.UUID,
    from_fy_year: int = Query(...),
    from_fy_month: int = Query(...),
    to_fy_year: int = Query(...),
    to_fy_month: int = Query(...),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_viewer),
):
    loc = await db.get(Location, location_id)
    if loc is None:
        raise HTTPException(status_code=404, detail="Location not found")

    periods = await _get_period_range(db, from_fy_year, from_fy_month, to_fy_year, to_fy_month)
    if not periods:
        return []

    period_ids = [p.id for p in periods]

    # Load target accounts
    rev_result = await db.execute(select(Account).where(Account.code == "REV-SALES"))
    rev_account = rev_result.scalar_one_or_none()

    dc_result = await db.execute(select(Account).where(Account.code.in_(DIRECT_COST_CODES)))
    dc_accounts = {a.code: a for a in dc_result.scalars().all()}

    revenue_account_ids = {rev_account.id} if rev_account else set()
    direct_cost_account_ids = {a.id for a in dc_accounts.values()}
    all_target_ids = revenue_account_ids | direct_cost_account_ids

    # Account mappings for location's entity
    entity_ids_for_mapping = set()
    if loc.entity_id:
        entity_ids_for_mapping.add(loc.entity_id)
    # Also grab all entity mappings for these target accounts
    mapping_result = await db.execute(
        select(AccountMapping).where(AccountMapping.target_account_id.in_(all_target_ids))
    )
    source_to_target: dict[tuple[uuid.UUID, str], tuple[uuid.UUID, float]] = {}
    for m in mapping_result.scalars().all():
        source_to_target[(m.entity_id, m.source_account_code)] = (m.target_account_id, float(m.multiplier))

    # Fetch je_lines for this location
    je_result = await db.execute(
        select(JeLine).where(
            JeLine.location_id == location_id,
            JeLine.period_id.in_(period_ids),
        )
    )

    # {period_id: {"revenue": float, "direct_costs": float}}
    period_data: dict[uuid.UUID, dict[str, float]] = defaultdict(lambda: {"revenue": 0.0, "direct_costs": 0.0})

    for jl in je_result.scalars().all():
        key = (jl.entity_id, jl.source_account_code)
        mapping = source_to_target.get(key)
        if not mapping:
            continue
        target_id, multiplier = mapping
        amount = float(jl.amount) * multiplier

        if target_id in revenue_account_ids:
            # Revenue is credit-normal (negative in DB) — negate to positive
            period_data[jl.period_id]["revenue"] += -amount
        elif target_id in direct_cost_account_ids:
            period_data[jl.period_id]["direct_costs"] += amount

    # Build points
    points: list[dict] = []
    for p in periods:
        data = period_data.get(p.id, {"revenue": 0.0, "direct_costs": 0.0})
        revenue = round(data["revenue"], 2)
        direct_costs = round(data["direct_costs"], 2)
        site_pl = round(revenue - direct_costs, 2)
        points.append({
            "period_label": _period_label(p),
            "fy_year": p.fy_year,
            "fy_month": p.fy_month,
            "revenue": revenue,
            "direct_costs": direct_costs,
            "site_pl": site_pl,
        })

    # Add MoM change on site_pl
    for i, pt in enumerate(points):
        if i == 0:
            pt["mom_change_pct"] = None
        else:
            prev = points[i - 1]["site_pl"]
            pt["mom_change_pct"] = round((pt["site_pl"] - prev) / prev * 100, 2) if prev != 0 else None

    return [LocationTimeSeriesPoint(**pt) for pt in points]


# ── POST /analytics/export ───────────────────────────────────────────────────


@router.post("/export")
async def export_analytics(
    payload: AnalyticsExportRequest,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_viewer),
):
    if payload.format != "xlsx":
        raise HTTPException(status_code=400, detail="Only xlsx format is currently supported")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    NAVY = "1F3D6E"
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    number_fmt = "#,##0"
    pct_fmt = "0.0%"
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    wb = Workbook()
    wb.remove(wb.active)

    if payload.report_type == "timeseries":
        await _export_timeseries(wb, db, payload.params, header_font, header_fill, number_fmt, pct_fmt, thin_border)
        filename = "kip_analytics_timeseries.xlsx"
    elif payload.report_type == "locations":
        await _export_locations(wb, db, payload.params, header_font, header_fill, number_fmt, pct_fmt, thin_border)
        filename = "kip_analytics_locations.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Invalid report_type")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _export_timeseries(wb, db, params, header_font, header_fill, number_fmt, pct_fmt, thin_border):
    from openpyxl.styles import Alignment

    metric = params.get("metric", "revenue")
    account_code = METRIC_CODES.get(metric, "REV-SALES")

    acct_result = await db.execute(select(Account).where(Account.code == account_code))
    account = acct_result.scalar_one_or_none()
    if not account:
        ws = wb.create_sheet(title="Time Series")
        ws.cell(row=1, column=1, value="Account not found")
        return

    from_fy_year = params.get("from_fy_year", 2024)
    from_fy_month = params.get("from_fy_month", 1)
    to_fy_year = params.get("to_fy_year", 2026)
    to_fy_month = params.get("to_fy_month", 12)

    periods = await _get_period_range(db, from_fy_year, from_fy_month, to_fy_year, to_fy_month)
    period_ids = [p.id for p in periods]

    q = select(ConsolidatedActual).where(
        ConsolidatedActual.period_id.in_(period_ids),
        ConsolidatedActual.account_id == account.id,
        ConsolidatedActual.is_group_total.is_(True),
    )
    result = await db.execute(q)
    sign = -1.0 if _NEGATE_METRICS else 1.0
    amount_by_period: dict[uuid.UUID, float] = defaultdict(float)
    for row in result.scalars().all():
        amount_by_period[row.period_id] += float(row.amount) * sign

    ws = wb.create_sheet(title=f"Time Series - {metric.upper()}")
    headers = ["Period", metric.upper(), "MoM %", "Rolling 3M", "Rolling 12M"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    for c in ["B", "C", "D", "E"]:
        ws.column_dimensions[c].width = 14

    points: list[dict] = []
    for p in periods:
        points.append({"period_label": _period_label(p), "value": amount_by_period.get(p.id, 0.0)})

    _add_mom_change(points)
    _add_rolling_averages(points)

    for i, pt in enumerate(points):
        row_idx = i + 2
        ws.cell(row=row_idx, column=1, value=pt["period_label"])
        ws.cell(row=row_idx, column=2, value=round(pt["value"], 2)).number_format = number_fmt
        if pt["mom_change_pct"] is not None:
            ws.cell(row=row_idx, column=3, value=round(pt["mom_change_pct"] / 100, 4)).number_format = pct_fmt
        if pt["rolling_3m_avg"] is not None:
            ws.cell(row=row_idx, column=4, value=round(pt["rolling_3m_avg"], 2)).number_format = number_fmt
        if pt["rolling_12m_avg"] is not None:
            ws.cell(row=row_idx, column=5, value=round(pt["rolling_12m_avg"], 2)).number_format = number_fmt


async def _export_locations(wb, db, params, header_font, header_fill, number_fmt, pct_fmt, thin_border):
    from openpyxl.styles import Alignment

    fy_year = params.get("fy_year", 2025)
    fy_month = params.get("fy_month")

    # Reuse the endpoint logic
    if fy_month:
        result = await db.execute(
            select(Period).where(Period.fy_year == fy_year, Period.fy_month == fy_month)
        )
    else:
        result = await db.execute(
            select(Period).where(Period.fy_year == fy_year).order_by(Period.fy_month)
        )
    target_periods = list(result.scalars().all())
    period_ids = [p.id for p in target_periods]

    loc_result = await db.execute(select(Location).where(Location.is_active.is_(True)))
    locations = {loc.id: loc for loc in loc_result.scalars().all()}

    ent_result = await db.execute(select(Entity))
    entities = {e.id: e for e in ent_result.scalars().all()}

    rev_result = await db.execute(select(Account).where(Account.code == "REV-SALES"))
    rev_account = rev_result.scalar_one_or_none()
    dc_result = await db.execute(select(Account).where(Account.code.in_(DIRECT_COST_CODES)))
    dc_accounts = {a.code: a for a in dc_result.scalars().all()}

    revenue_account_ids = {rev_account.id} if rev_account else set()
    direct_cost_account_ids = {a.id for a in dc_accounts.values()}
    all_target_ids = revenue_account_ids | direct_cost_account_ids

    mapping_result = await db.execute(
        select(AccountMapping).where(AccountMapping.target_account_id.in_(all_target_ids))
    )
    source_to_target: dict[tuple[uuid.UUID, str], tuple[uuid.UUID, float]] = {}
    for m in mapping_result.scalars().all():
        source_to_target[(m.entity_id, m.source_account_code)] = (m.target_account_id, float(m.multiplier))

    je_result = await db.execute(
        select(JeLine).where(JeLine.period_id.in_(period_ids), JeLine.location_id.isnot(None))
    )
    loc_actuals: dict[uuid.UUID, dict[str, float]] = defaultdict(lambda: {"revenue": 0.0, "direct_costs": 0.0})
    for jl in je_result.scalars().all():
        key = (jl.entity_id, jl.source_account_code)
        mapping = source_to_target.get(key)
        if not mapping:
            continue
        target_id, multiplier = mapping
        amount = float(jl.amount) * multiplier
        if target_id in revenue_account_ids:
            loc_actuals[jl.location_id]["revenue"] += -amount
        elif target_id in direct_cost_account_ids:
            loc_actuals[jl.location_id]["direct_costs"] += amount

    ws = wb.create_sheet(title="Location Performance")
    headers = ["Site", "State", "Entity", "Revenue", "Direct Costs", "Site P&L"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    for c in ["D", "E", "F"]:
        ws.column_dimensions[c].width = 14

    row_idx = 2
    sorted_locs = sorted(loc_actuals.items(), key=lambda x: x[1]["revenue"] - x[1]["direct_costs"], reverse=True)
    for loc_id, data in sorted_locs:
        loc = locations.get(loc_id)
        if not loc:
            continue
        entity = entities.get(loc.entity_id) if loc.entity_id else None
        revenue = round(data["revenue"], 2)
        direct_costs = round(data["direct_costs"], 2)
        site_pl = round(revenue - direct_costs, 2)

        ws.cell(row=row_idx, column=1, value=loc.name or loc.code)
        ws.cell(row=row_idx, column=2, value=loc.state or "")
        ws.cell(row=row_idx, column=3, value=entity.code if entity else "")
        ws.cell(row=row_idx, column=4, value=revenue).number_format = number_fmt
        ws.cell(row=row_idx, column=5, value=direct_costs).number_format = number_fmt
        ws.cell(row=row_idx, column=6, value=site_pl).number_format = number_fmt

        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1
